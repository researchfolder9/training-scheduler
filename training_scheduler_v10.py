import streamlit as st
import random
import calendar
import io
from datetime import date
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

st.set_page_config(page_title="Training Scheduler", layout="wide")

DEFAULT_CLASSES = [
    {"name": "Mech / Elec Torque",                   "duration": 4.0,  "all_day": False, "room_restriction": None,                   "priority": False},
    {"name": "Safety Wire / Cable Installation",      "duration": 3.0,  "all_day": False, "room_restriction": None,                   "priority": False},
    {"name": "Smart Torque",                          "duration": 2.0,  "all_day": False, "room_restriction": None,                   "priority": True},
    {"name": "Threaded Insert Installation",          "duration": 3.5,  "all_day": False, "room_restriction": None,                   "priority": False},
    {"name": "Fluid Fittings Installation",           "duration": 2.5,  "all_day": False, "room_restriction": None,                   "priority": False},
    {"name": "Wire Harness Mate / Demate",            "duration": 3.0,  "all_day": False, "room_restriction": None,                   "priority": False},
    {"name": "Wire Harness Routing and Installation", "duration": 4.0,  "all_day": False, "room_restriction": None,                   "priority": False},
    {"name": "Conversion Coating",                    "duration": 1.5,  "all_day": False, "room_restriction": None,                   "priority": False},
    {"name": "Application of Sealants",               "duration": 3.0,  "all_day": False, "room_restriction": None,                   "priority": False},
    {"name": "Component Adhesive Bonding",            "duration": 2.0,  "all_day": False, "room_restriction": None,                   "priority": False},
    {"name": "Bonding Structural",                    "duration": 4.0,  "all_day": False, "room_restriction": None,                   "priority": False},
    {"name": "MPS Liquid Shim",                       "duration": 2.0,  "all_day": False, "room_restriction": None,                   "priority": False},
    {"name": "Confined Space",                        "duration": 3.0,  "all_day": False, "room_restriction": None,                   "priority": True},
    {"name": "Strain Gauge Installation",             "duration": 4.0,  "all_day": False, "room_restriction": None,                   "priority": False},
    {"name": "Lock Out / Tag Out",                    "duration": 2.0,  "all_day": False, "room_restriction": None,                   "priority": True},
    {"name": "Rynglok - Axial Swage",                 "duration": 2.0,  "all_day": False, "room_restriction": None,                   "priority": False},
    {"name": "IPC 620",                               "duration": 6.0,  "all_day": True,  "room_restriction": "Avionics Training Lab", "priority": False},
    {"name": "J-STD",                                 "duration": 6.0,  "all_day": True,  "room_restriction": "Avionics Training Lab", "priority": False},
]

SLOT1_MIN, SLOT1_MAX = 3.0, 4.0
SLOT2_MIN, SLOT2_MAX = 1.5, 2.5

DEFAULT_SHIFTS = {
    "A": {"label": "A-Shift", "start": "06:30", "end": "14:00", "days": ["Monday","Tuesday","Wednesday","Thursday"]},
    "B": {"label": "B-Shift", "start": "16:00", "end": "02:00", "days": ["Monday","Tuesday","Wednesday","Thursday"]},
    "C": {"label": "C-Shift", "start": "06:30", "end": "14:00", "days": ["Friday","Saturday","Sunday"]},
}

DEFAULT_INSTRUCTORS = [
    {"name": "Katie",            "shift": "A"},
    {"name": "Chris",            "shift": "A"},
    {"name": "Eric",             "shift": "A"},
    {"name": "Dave",             "shift": "B"},
    {"name": "Tempi Placeholda", "shift": "C"},
]

INSTRUCTOR_COLORS = {
    "Katie":            {"bg": "#166534", "text": "#ffffff"},
    "Chris":            {"bg": "#1d4ed8", "text": "#ffffff"},
    "Eric":             {"bg": "#c2410c", "text": "#ffffff"},
    "Dave":             {"bg": "#490b6b", "text": "#ffffff"},
    "Tempi Placeholda": {"bg": "#0f766e", "text": "#ffffff"},
}
DEFAULT_COLOR = {"bg": "#444444", "text": "#ffffff"}

DEFAULT_ROOMS = ["Avionics Training Lab", "Galileo", "Newton", "Classroom C"]
PRIORITY_DEFAULT = 5
STANDARD_DEFAULT = 1
MEAL_MINS = 30
PREP_MINS = 30


def init_state():
    if "step" not in st.session_state:
        st.session_state.step = 1
    if "schedule_month" not in st.session_state:
        st.session_state.schedule_month = None
    if "schedule_year" not in st.session_state:
        st.session_state.schedule_year = None
    if "class_requirements" not in st.session_state:
        st.session_state.class_requirements = {
            c["name"]: (PRIORITY_DEFAULT if c["priority"] else STANDARD_DEFAULT)
            for c in DEFAULT_CLASSES
        }
    if "constraints" not in st.session_state:
        st.session_state.constraints = {"holidays": [], "meetings": [], "pto": []}
    if "qualifications" not in st.session_state:
        all_cn = [c["name"] for c in DEFAULT_CLASSES]
        st.session_state.qualifications = {
            inst["name"]: {cn: True for cn in all_cn}
            for inst in DEFAULT_INSTRUCTORS
        }
    if "generated_schedule" not in st.session_state:
        st.session_state.generated_schedule = None

init_state()

STEPS = ["1 - Month & Year","2 - Course Requirements","3 - Constraints","4 - Instructor Review","5 - Generate"]
st.sidebar.title("Training Scheduler")
st.sidebar.markdown("---")
for idx, label in enumerate(STEPS, 1):
    prefix = "OK " if st.session_state.step > idx else (">> " if st.session_state.step == idx else "   ")
    if st.sidebar.button(prefix + label, key="nav_"+str(idx), use_container_width=True):
        st.session_state.step = idx
if st.session_state.generated_schedule:
    st.sidebar.markdown("---")
    if st.sidebar.button("Calendar View", key="nav_cal", use_container_width=True):
        st.session_state.step = 6
    if st.sidebar.button("Day Detail View", key="nav_day", use_container_width=True):
        st.session_state.step = 7
st.sidebar.markdown("---")
st.sidebar.caption("Complete each step in order before generating.")

def next_step(): st.session_state.step = min(st.session_state.step + 1, 5)
def prev_step(): st.session_state.step = max(st.session_state.step - 1, 1)

def get_month_dates(year, month):
    _, num_days = calendar.monthrange(year, month)
    return [date(year, month, d) for d in range(1, num_days + 1)]

def day_name(d): return d.strftime("%A")

def time_to_minutes(t_str):
    h, m = map(int, t_str.split(":"))
    return h * 60 + m

def minutes_to_time(mins):
    mins = int(mins) % (24 * 60)
    return str(mins // 60).zfill(2) + ":" + str(mins % 60).zfill(2)

def inst_color(name):
    return INSTRUCTOR_COLORS.get(name, DEFAULT_COLOR)

def get_shift_window(sk):
    sh = DEFAULT_SHIFTS[sk]
    s = time_to_minutes(sh["start"])
    e = time_to_minutes(sh["end"])
    if e <= s: e += 24 * 60
    return s, e

def inst_blocked(iname, dobj, bstart, bend):
    ds = dobj.isoformat()
    for h in st.session_state.constraints["holidays"]:
        if h["date"] == ds: return True
    for p in st.session_state.constraints["pto"]:
        if p["instructor"] == iname and p["date"] == ds: return True
    for m in st.session_state.constraints["meetings"]:
        if m["date"] == ds and iname in m["instructors"]:
            ms = time_to_minutes(m["start"])
            me = ms + int(m["duration_hrs"] * 60)
            if not (bend <= ms or bstart >= me):
                return True
    return False

def is_day_blocked(iname, dobj):
    ds = dobj.isoformat()
    for h in st.session_state.constraints["holidays"]:
        if h["date"] == ds: return True
    for p in st.session_state.constraints["pto"]:
        if p["instructor"] == iname and p["date"] == ds: return True
    return False

def find_room(course, dobj, cs, ce, rsched):
    if course.get("room_restriction"):
        rooms = [course["room_restriction"]]
    else:
        rooms = [r for r in DEFAULT_ROOMS if r != "Avionics Training Lab"]
    for room in rooms:
        busy = rsched.get((room, dobj.isoformat()), [])
        if not any(not (ce <= rs or cs >= re) for rs, re, _ in busy):
            return room
    return None

def schedule_two_class_day(sk, dobj, iname, c1_name, c2_name, cmap, rsched, isched):
    sh_s, sh_e = get_shift_window(sk)
    c1 = cmap[c1_name]
    dur1 = int(c1["duration"] * 60)

    if c1["all_day"]:
        ps1 = sh_s; cs1 = ps1 + PREP_MINS; ce1 = cs1 + dur1 + MEAL_MINS
        if ce1 > sh_e: return None, None, None
        if inst_blocked(iname, dobj, cs1, ce1): return None, None, None
        room1 = find_room(c1, dobj, cs1, ce1, rsched)
        if room1 is None: return None, None, None
        return _make_session(sk, dobj, iname, c1_name, c1, ps1, cs1, ce1, room1), None, None

    ps1 = sh_s; cs1 = ps1 + PREP_MINS; ce1 = cs1 + dur1
    if ce1 + MEAL_MINS > sh_e: return None, None, None
    if inst_blocked(iname, dobj, cs1, ce1): return None, None, None
    room1 = find_room(c1, dobj, cs1, ce1, rsched)
    if room1 is None: return None, None, None

    meal_start = ce1
    s1 = _make_session(sk, dobj, iname, c1_name, c1, ps1, cs1, ce1, room1)

    if c2_name is None:
        return s1, None, meal_start

    c2 = cmap[c2_name]
    dur2 = int(c2["duration"] * 60)
    ps2 = meal_start + MEAL_MINS; cs2 = ps2 + PREP_MINS; ce2 = cs2 + dur2
    if ce2 > sh_e: return s1, None, meal_start
    if inst_blocked(iname, dobj, cs2, ce2): return s1, None, meal_start
    room2 = find_room(c2, dobj, cs2, ce2, rsched)
    if room2 is None: return s1, None, meal_start

    return s1, _make_session(sk, dobj, iname, c2_name, c2, ps2, cs2, ce2, room2), meal_start

def _make_session(sk, dobj, iname, cn, course, ps, cs, ce, room):
    return {
        "date": dobj.isoformat(), "shift": sk, "course": cn,
        "instructor": iname, "room": room,
        "prep_start": minutes_to_time(ps), "class_start": minutes_to_time(cs), "class_end": minutes_to_time(ce),
        "prep_start_min": ps, "class_start_min": cs, "class_end_min": ce,
        "duration_hrs": course["duration"], "all_day": course["all_day"],
    }

def commit_session(s, rsched, isched, icount, scc):
    iname, date_iso, sk, cn = s["instructor"], s["date"], s["shift"], s["course"]
    ki = (iname, date_iso)
    isched.setdefault(ki, []).append((s["prep_start_min"], s["class_end_min"], cn))
    isched[ki].sort(key=lambda x: x[0])
    rsched.setdefault((s["room"], date_iso), []).append((s["class_start_min"], s["class_end_min"], cn))
    rsched[(s["room"], date_iso)].sort(key=lambda x: x[0])
    icount[iname] = icount.get(iname, 0) + 1
    scc[sk][cn] = scc[sk].get(cn, 0) + 1

def pick_slot1(pool, cmap, scc, sk, reqs, used_cn=None):
    def shift_over(cn): return scc[sk][cn] >= max(1, reqs.get(cn, 0))
    def total_rem(cn): return max(0, reqs.get(cn, 0) - sum(scc[s][cn] for s in ["A","B","C"]))
    candidates = [cn for cn in pool if cn != used_cn]
    for cn in candidates:
        if SLOT1_MIN <= cmap[cn]["duration"] <= SLOT1_MAX and not shift_over(cn) and total_rem(cn) > 0:
            return cn
    for cn in candidates:
        if SLOT1_MIN <= cmap[cn]["duration"] <= SLOT1_MAX and not shift_over(cn):
            return cn
    for cn in candidates:
        if not shift_over(cn):
            return cn
    return candidates[0] if candidates else None

def pick_slot2(pool, cmap, scc, sk, reqs, exclude_cn=None):
    def total_rem(cn): return max(0, reqs.get(cn, 0) - sum(scc[s][cn] for s in ["A","B","C"]))
    def shift_over(cn): return scc[sk][cn] >= max(1, reqs.get(cn, 0))
    pool_f = [cn for cn in pool if cn != exclude_cn]
    # 1. In range, needed, not over
    best = [cn for cn in pool_f if SLOT2_MIN <= cmap[cn]["duration"] <= SLOT2_MAX and total_rem(cn) > 0 and not shift_over(cn)]
    if best: best.sort(key=lambda cn: scc[sk][cn]); return best[0]
    # 2. In range, not over
    ok = [cn for cn in pool_f if SLOT2_MIN <= cmap[cn]["duration"] <= SLOT2_MAX and not shift_over(cn)]
    if ok: ok.sort(key=lambda cn: scc[sk][cn]); return ok[0]
    return None

def qualified_pool_for(iname, sk, scc, reqs, cmap):
    all_cn = [c["name"] for c in DEFAULT_CLASSES]
    pool = [cn for cn in all_cn if not cmap[cn]["all_day"] and st.session_state.qualifications[iname].get(cn, True)]
    def total_rem(cn): return max(0, reqs.get(cn, 0) - sum(scc[s][cn] for s in ["A","B","C"]))
    pool.sort(key=lambda cn: (-total_rem(cn), scc[sk][cn]))
    return pool

def generate_schedule():
    month = st.session_state.schedule_month
    year = st.session_state.schedule_year
    month_dates = get_month_dates(year, month)
    reqs = dict(st.session_state.class_requirements)
    quals = st.session_state.qualifications
    sessions = []
    flags = []
    rsched = {}
    isched = {}
    meal_map = {}
    all_cn = [c["name"] for c in DEFAULT_CLASSES]
    cmap = {c["name"]: c for c in DEFAULT_CLASSES}
    shift_days = {"A":["Monday","Tuesday","Wednesday","Thursday"],
                  "B":["Monday","Tuesday","Wednesday","Thursday"],
                  "C":["Friday","Saturday","Sunday"]}
    ibs = {"A":[i for i in DEFAULT_INSTRUCTORS if i["shift"]=="A"],
           "B":[i for i in DEFAULT_INSTRUCTORS if i["shift"]=="B"],
           "C":[i for i in DEFAULT_INSTRUCTORS if i["shift"]=="C"]}
    scc = {s:{cn:0 for cn in all_cn} for s in ["A","B","C"]}
    icount = {i["name"]:0 for i in DEFAULT_INSTRUCTORS}
    sdates = {sk:[d for d in month_dates if day_name(d) in shift_days[sk]] for sk in ["A","B","C"]}
    used_slots = set()

    def try_assign_forced(iname, sk, dobj, force_c1):
        if (iname, dobj.isoformat()) in used_slots: return False
        if is_day_blocked(iname, dobj): return False
        pool = qualified_pool_for(iname, sk, scc, reqs, cmap)
        c2 = pick_slot2(pool, cmap, scc, sk, reqs, exclude_cn=force_c1)
        s1, s2, meal_start = schedule_two_class_day(sk, dobj, iname, force_c1, c2, cmap, rsched, isched)
        if s1 is None:
            s1, s2, meal_start = schedule_two_class_day(sk, dobj, iname, force_c1, None, cmap, rsched, isched)
        if s1 is None: return False
        commit_session(s1, rsched, isched, icount, scc); sessions.append(s1)
        if s2: commit_session(s2, rsched, isched, icount, scc); sessions.append(s2)
        if meal_start is not None: meal_map[(iname, dobj.isoformat())] = meal_start
        used_slots.add((iname, dobj.isoformat()))
        return True

    def try_assign_day(iname, sk, dobj):
        if (iname, dobj.isoformat()) in used_slots: return False
        if is_day_blocked(iname, dobj): return False
        pool = qualified_pool_for(iname, sk, scc, reqs, cmap)
        allday = [cn for cn in all_cn if cmap[cn]["all_day"] and quals[iname].get(cn, True)]
        for cn in allday:
            if max(0, reqs.get(cn,0) - sum(scc[s][cn] for s in ["A","B","C"])) > 0:
                s1, _, _ = schedule_two_class_day(sk, dobj, iname, cn, None, cmap, rsched, isched)
                if s1:
                    commit_session(s1, rsched, isched, icount, scc); sessions.append(s1)
                    used_slots.add((iname, dobj.isoformat())); return True
        c1 = pick_slot1(pool, cmap, scc, sk, reqs)
        if c1 is None: return False
        c2 = pick_slot2(pool, cmap, scc, sk, reqs, exclude_cn=c1)
        s1, s2, meal_start = schedule_two_class_day(sk, dobj, iname, c1, c2, cmap, rsched, isched)
        if s1 is None:
            s1, s2, meal_start = schedule_two_class_day(sk, dobj, iname, c1, None, cmap, rsched, isched)
        if s1 is None:
            for cn in pool:
                s1, s2, meal_start = schedule_two_class_day(sk, dobj, iname, cn, None, cmap, rsched, isched)
                if s1: break
        if s1 is None: return False
        commit_session(s1, rsched, isched, icount, scc); sessions.append(s1)
        if s2: commit_session(s2, rsched, isched, icount, scc); sessions.append(s2)
        if meal_start is not None: meal_map[(iname, dobj.isoformat())] = meal_start
        used_slots.add((iname, dobj.isoformat()))
        return True

    # Phase 1: guarantee each non-allday course at least once per shift
    for sk in ["A","B","C"]:
        dlist = sdates[sk]
        if not dlist: continue
        for course in DEFAULT_CLASSES:
            if course["all_day"]: continue
            cn = course["name"]
            if scc[sk][cn] > 0: continue
            qualified = [i for i in ibs[sk] if quals[i["name"]].get(cn, True)]
            if not qualified:
                flags.append(cn + " has no qualified instructor on " + DEFAULT_SHIFTS[sk]["label"] + " — skipped.")
                continue
            placed = False
            for dobj in dlist:
                if placed: break
                for inst in sorted(qualified, key=lambda i: icount[i["name"]]):
                    if (inst["name"], dobj.isoformat()) in used_slots: continue
                    if try_assign_forced(inst["name"], sk, dobj, cn):
                        placed = True; break
            if not placed:
                flags.append("Could not schedule " + cn + " on " + DEFAULT_SHIFTS[sk]["label"] + " (minimum once).")

    # Phase 1b: all-day courses minimum once per qualifying shift
    for sk in ["A","B","C"]:
        dlist = sdates[sk]
        if not dlist: continue
        for course in DEFAULT_CLASSES:
            if not course["all_day"]: continue
            cn = course["name"]
            if scc[sk][cn] > 0: continue
            qualified = [i for i in ibs[sk] if quals[i["name"]].get(cn, True)]
            if not qualified: continue
            placed = False
            for dobj in dlist:
                if placed: break
                for inst in sorted(qualified, key=lambda i: icount[i["name"]]):
                    if (inst["name"], dobj.isoformat()) in used_slots: continue
                    s1, _, _ = schedule_two_class_day(sk, dobj, inst["name"], cn, None, cmap, rsched, isched)
                    if s1:
                        commit_session(s1, rsched, isched, icount, scc); sessions.append(s1)
                        used_slots.add((inst["name"], dobj.isoformat())); placed = True; break

    # Phase 2: fill required counts, rotating across shifts
    for cn in all_cn:
        req = reqs.get(cn, 0)
        need = req - sum(scc[s][cn] for s in ["A","B","C"])
        if need <= 0: continue
        for sk in ["A","B","C","A","B","C"]:
            if need <= 0: break
            dlist = list(sdates[sk]); random.shuffle(dlist)
            for dobj in dlist:
                if need <= 0: break
                qualified = [i for i in ibs[sk] if quals[i["name"]].get(cn, True)]
                for inst in sorted(qualified, key=lambda i: icount[i["name"]]):
                    if (inst["name"], dobj.isoformat()) in used_slots: continue
                    if try_assign_forced(inst["name"], sk, dobj, cn):
                        need -= 1; break
        if need > 0:
            flags.append("Could only schedule " + cn + " " + str(reqs.get(cn,0)-need) + "/" + str(reqs.get(cn,0)) + " times.")

    # Phase 3: fill remaining open instructor/day slots
    for sk in ["A","B","C"]:
        dlist = list(sdates[sk]); random.shuffle(dlist)
        for dobj in dlist:
            for inst in sorted(ibs[sk], key=lambda i: icount[i["name"]]):
                try_assign_day(inst["name"], sk, dobj)

    # ── Phase 4: Gap Fill Pass ─────────────────────────────────────────────────
    # For every instructor/day already in used_slots, check if there's a free
    # window after their last class + meal break. If so, add one more course,
    # always picking the globally least-scheduled qualified course first.
    def gap_fill_pass():
        inst_map = {i["name"]: i for i in DEFAULT_INSTRUCTORS}
        # Sort used slots so we process them deterministically
        for (iname, date_iso) in sorted(used_slots):
            dobj = date.fromisoformat(date_iso)
            sk = inst_map[iname]["shift"]
            _, sh_e = get_shift_window(sk)

            # Find the end of the last committed block for this instructor/day
            day_blocks = isched.get((iname, date_iso), [])
            if not day_blocks: continue

            last_end = max(b[1] for b in day_blocks)  # class_end of last session

            # Need: MEAL (if not already had one) + PREP + class
            # Use a simple rule: free_start = last_end + MEAL_MINS + PREP_MINS
            free_start = last_end + MEAL_MINS + PREP_MINS

            # Check if there's space for at least the shortest course (1.5hr = 90 min)
            if free_start + 90 > sh_e: continue

            available_mins = sh_e - free_start

            # Build pool: qualified, non-allday, fits in available time
            pool = [c["name"] for c in DEFAULT_CLASSES
                    if not c["all_day"]
                    and quals[iname].get(c["name"], True)
                    and int(c["duration"] * 60) <= available_mins]
            if not pool: continue

            # Sort by globally least-scheduled first (even spread), then least on this shift
            def global_count(cn):
                return sum(scc[s][cn] for s in ["A","B","C"])
            pool.sort(key=lambda cn: (global_count(cn), scc[sk][cn]))

            # Try each candidate until one fits (room availability etc.)
            for cn in pool:
                course = cmap[cn]
                dur = int(course["duration"] * 60)
                cs = free_start
                ce = cs + dur
                if ce > sh_e: continue
                if inst_blocked(iname, dobj, cs, ce): continue
                room = find_room(course, dobj, cs, ce, rsched)
                if room is None: continue

                # Use last_end as the "prep start" visually (meal break fills that gap)
                ps = last_end + MEAL_MINS  # meal, then this is prep
                s = _make_session(sk, dobj, iname, cn, course, ps, cs, ce, room)
                commit_session(s, rsched, isched, icount, scc)
                sessions.append(s)
                # Record meal break if not already present
                if (iname, date_iso) not in meal_map:
                    meal_map[(iname, date_iso)] = last_end
                break  # one extra class per gap fill pass per instructor/day

    # Run gap fill up to 3 times to catch cascading gaps (e.g. very short slot1)
    for _ in range(3):
        gap_fill_pass()
    # ──────────────────────────────────────────────────────────────────────────

    return {"sessions": sessions, "flags": flags, "meal_map": meal_map}


def step1():
    st.title("Step 1 - Select Month & Year")
    col1, col2 = st.columns(2)
    with col1:
        year = st.selectbox("Year", list(range(2025,2030)),
            index=list(range(2025,2030)).index(st.session_state.schedule_year or 2026))
    with col2:
        mnames = list(calendar.month_name)[1:]
        midx = (st.session_state.schedule_month or 3) - 1
        month_sel = st.selectbox("Month", mnames, index=midx)
        month_num = mnames.index(month_sel) + 1
    st.markdown("---")
    st.subheader("Calendar Preview")
    sdmap = {"Monday":"A+B","Tuesday":"A+B","Wednesday":"A+B","Thursday":"A+B",
             "Friday":"C","Saturday":"C","Sunday":"C"}
    dates = get_month_dates(year, month_num)
    weeks, week = [], []
    for _ in range(dates[0].weekday()): week.append(None)
    for d in dates:
        week.append(d)
        if len(week)==7: weeks.append(week); week=[]
    if week:
        while len(week)<7: week.append(None)
        weeks.append(week)
    hcols = st.columns(7)
    for i, h in enumerate(["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]):
        hcols[i].markdown("**"+h+"**")
    for wk in weeks:
        cols = st.columns(7)
        for i, d in enumerate(wk):
            if d is None: cols[i].write(" ")
            else: cols[i].markdown("**"+str(d.day)+"** "+sdmap.get(day_name(d),"Off"))
    st.markdown("---")
    if st.button("Next: Course Requirements", type="primary"):
        st.session_state.schedule_year = year
        st.session_state.schedule_month = month_num
        next_step()

def step2():
    st.title("Step 2 - Course Requirements")
    for c in DEFAULT_CLASSES:
        default = PRIORITY_DEFAULT if c["priority"] else STANDARD_DEFAULT
        cur = st.session_state.class_requirements.get(c["name"], default)
        tag = " [PRIORITY]" if c["priority"] else ""
        aday = " (All Day - Avionics Lab)" if c["all_day"] else ""
        val = st.number_input(c["name"]+tag+" - "+str(c["duration"])+"hrs"+aday,
                              min_value=0, max_value=100, value=cur, step=1, key="req_"+c["name"])
        st.session_state.class_requirements[c["name"]] = val
    st.markdown("---")
    st.info("Total required sessions this month: "+str(sum(st.session_state.class_requirements.values())))
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Back"): prev_step()
    with col2:
        if st.button("Next: Constraints", type="primary"): next_step()

def step3():
    st.title("Step 3 - Constraints")
    inst_names = [i["name"] for i in DEFAULT_INSTRUCTORS]
    month = st.session_state.schedule_month or 3
    year = st.session_state.schedule_year or 2026
    month_dates = get_month_dates(year, month)
    date_options = [d.strftime("%A, %B %d") for d in month_dates]
    date_lookup = {d.strftime("%A, %B %d"): d for d in month_dates}
    st.subheader("Holidays / Blackout Days")
    with st.form("hform", clear_on_submit=True):
        hc1, hc2 = st.columns([3,1])
        hdate = hc1.selectbox("Date", date_options, key="hdate")
        hlabel = hc1.text_input("Label", key="hlabel")
        hc2.markdown("<br><br><br>", unsafe_allow_html=True)
        if hc2.form_submit_button("Add") and hlabel:
            st.session_state.constraints["holidays"].append({"date": date_lookup[hdate].isoformat(), "label": hlabel})
    for i, h in enumerate(st.session_state.constraints["holidays"]):
        c1, c2 = st.columns([5,1])
        c1.markdown("Holiday: **"+h["label"]+"** on "+h["date"])
        if c2.button("Remove", key="dh_"+str(i)): st.session_state.constraints["holidays"].pop(i); st.rerun()
    st.markdown("---")
    st.subheader("Instructor PTO")
    with st.form("pform", clear_on_submit=True):
        pc1, pc2, pc3 = st.columns(3)
        pto_inst = pc1.selectbox("Instructor", inst_names, key="ptoi")
        pto_date = pc2.selectbox("Date", date_options, key="ptod")
        pc3.markdown("<br><br>", unsafe_allow_html=True)
        if pc3.form_submit_button("Add PTO"):
            st.session_state.constraints["pto"].append({"instructor": pto_inst, "date": date_lookup[pto_date].isoformat()})
    for i, p in enumerate(st.session_state.constraints["pto"]):
        c1, c2 = st.columns([5,1])
        c1.markdown("PTO: **"+p["instructor"]+"** on "+p["date"])
        if c2.button("Remove", key="dp_"+str(i)): st.session_state.constraints["pto"].pop(i); st.rerun()
    st.markdown("---")
    st.subheader("Meetings")
    with st.form("mform", clear_on_submit=True):
        mc1, mc2 = st.columns(2)
        mdate = mc1.selectbox("Date", date_options, key="mdate")
        minst = mc1.multiselect("Instructors", inst_names, key="minst")
        mlabel = mc1.text_input("Label", value="Team Meeting", key="mlabel")
        mstart = mc2.time_input("Start Time", key="mstart")
        mdur = mc2.number_input("Duration (hrs)", min_value=0.5, max_value=8.0, value=1.0, step=0.5, key="mdur")
        if st.form_submit_button("Add Meeting") and minst and mstart:
            st.session_state.constraints["meetings"].append({
                "date": date_lookup[mdate].isoformat(), "instructors": minst,
                "label": mlabel, "start": mstart.strftime("%H:%M"), "duration_hrs": float(mdur)})
    for i, m in enumerate(st.session_state.constraints["meetings"]):
        c1, c2 = st.columns([5,1])
        c1.markdown("Meeting: **"+m["label"]+"** on "+m["date"]+" at "+m["start"]+
                    " ("+str(m["duration_hrs"])+"hr) - "+", ".join(m["instructors"]))
        if c2.button("Remove", key="dm_"+str(i)): st.session_state.constraints["meetings"].pop(i); st.rerun()
    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Back"): prev_step()
    with col2:
        if st.button("Next: Instructor Review", type="primary"): next_step()

def step4():
    st.title("Step 4 - Instructor Review & Qualifications")
    all_cn = [c["name"] for c in DEFAULT_CLASSES]
    for inst in DEFAULT_INSTRUCTORS:
        name = inst["name"]
        sh = DEFAULT_SHIFTS[inst["shift"]]
        with st.expander(name+" - "+sh["label"]+" ("+sh["start"]+" to "+sh["end"]+", "+", ".join(sh["days"])+")", expanded=False):
            cols = st.columns(3)
            for j, cn in enumerate(all_cn):
                cur = st.session_state.qualifications[name].get(cn, True)
                nv = cols[j%3].checkbox(cn, value=cur, key="q_"+name+"_"+cn)
                st.session_state.qualifications[name][cn] = nv
    st.markdown("---")
    tbl = "| Instructor | Shift | Qualified Courses |\n|---|---|---|\n"
    for inst in DEFAULT_INSTRUCTORS:
        name = inst["name"]
        qc = sum(1 for v in st.session_state.qualifications[name].values() if v)
        tbl += "| "+name+" | "+DEFAULT_SHIFTS[inst["shift"]]["label"]+" | "+str(qc)+" / "+str(len(DEFAULT_CLASSES))+" |\n"
    st.markdown(tbl)
    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Back"): prev_step()
    with col2:
        if st.button("Next: Generate", type="primary"): next_step()

def step5():
    st.title("Step 5 - Review & Generate")
    month = st.session_state.schedule_month
    year = st.session_state.schedule_year
    if not month or not year:
        st.warning("Please go back to Step 1 and select a month and year.")
        if st.button("Back to Step 1"): st.session_state.step = 1
        return
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Schedule Period")
        st.markdown("**"+calendar.month_name[month]+" "+str(year)+"**")
        st.subheader("Course Requirements")
        for c in DEFAULT_CLASSES:
            count = st.session_state.class_requirements.get(c["name"], STANDARD_DEFAULT)
            tag = " [P]" if c["priority"] else ""
            st.markdown("- "+c["name"]+tag+": **"+str(count)+"**")
    with col2:
        st.subheader("Constraints")
        h = st.session_state.constraints["holidays"]
        p = st.session_state.constraints["pto"]
        m = st.session_state.constraints["meetings"]
        st.markdown("Holidays: "+str(len(h))+" | PTO: "+str(len(p))+" | Meetings: "+str(len(m)))
        st.subheader("Instructor Qualifications")
        for inst in DEFAULT_INSTRUCTORS:
            name = inst["name"]
            qc = sum(1 for v in st.session_state.qualifications[name].values() if v)
            st.markdown("- "+name+" ("+DEFAULT_SHIFTS[inst["shift"]]["label"]+"): "+str(qc)+"/"+str(len(DEFAULT_CLASSES)))
    st.markdown("---")
    st.info("Total minimum required sessions: "+str(sum(st.session_state.class_requirements.values())))
    if st.button("GENERATE SCHEDULE", type="primary", use_container_width=True):
        with st.spinner("Building schedule..."):
            result = generate_schedule()
            st.session_state.generated_schedule = result
        st.success("Schedule generated! Use the sidebar to view Calendar or Day Detail.")
        st.rerun()
    if st.session_state.generated_schedule:
        sched = st.session_state.generated_schedule
        sessions = sched["sessions"]
        if sched["flags"]:
            st.subheader("Scheduling Flags")
            for f in sched["flags"]: st.warning(f)
        st.markdown("---")
        st.subheader("Quick Summary")
        st.markdown("#### Instructor Load")
        inst_html = ("<table style='width:100%;border-collapse:collapse;font-size:13px'>"
            "<tr style='background:#222;color:#fff'>"
            "<th style='padding:6px 10px;text-align:left'>Instructor</th>"
            "<th style='padding:6px 10px;text-align:left'>Shift</th>"
            "<th style='padding:6px 10px;text-align:center'>Sessions</th>"
            "<th style='padding:6px 10px;text-align:center'>Hours</th></tr>")
        for idx, inst in enumerate(DEFAULT_INSTRUCTORS):
            name = inst["name"]
            col = inst_color(name)
            inst_s = [s for s in sessions if s["instructor"]==name]
            total_hrs = sum(s["duration_hrs"] for s in inst_s)
            row_bg = "#f4f4f4" if idx%2==0 else "#ffffff"
            inst_html += ("<tr style='background:"+row_bg+"'>"
                "<td style='padding:6px 10px'><span style='background:"+col["bg"]+";color:"+col["text"]+
                ";padding:2px 8px;border-radius:4px;font-weight:bold'>"+name+"</span></td>"
                "<td style='padding:6px 10px'>"+DEFAULT_SHIFTS[inst["shift"]]["label"]+"</td>"
                "<td style='padding:6px 10px;text-align:center;font-weight:bold'>"+str(len(inst_s))+"</td>"
                "<td style='padding:6px 10px;text-align:center;font-weight:bold'>"+str(round(total_hrs,1))+" hrs</td></tr>")
        inst_html += "</table>"
        st.markdown(inst_html, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("#### Course Offerings by Shift")
        st.markdown("<small>"
            "<span style='background:#d4edda;padding:2px 6px;border-radius:3px'>Green</span> = met &nbsp;|&nbsp;"
            "<span style='background:#fff3cd;padding:2px 6px;border-radius:3px'>Yellow</span> = partial &nbsp;|&nbsp;"
            "<span style='background:#f8d7da;padding:2px 6px;border-radius:3px'>Red</span> = none &nbsp;|&nbsp; ★ = Priority"
            "</small>", unsafe_allow_html=True)
        tbl_html = ("<table style='width:100%;border-collapse:collapse;font-size:12px;margin-top:6px'>"
            "<tr style='background:#222;color:#fff'>"
            "<th style='padding:5px 8px;text-align:left'>Course</th>"
            "<th style='padding:5px 8px;text-align:center'>A-Shift</th>"
            "<th style='padding:5px 8px;text-align:center'>B-Shift</th>"
            "<th style='padding:5px 8px;text-align:center'>C-Shift</th>"
            "<th style='padding:5px 8px;text-align:center'>Total / Req</th></tr>")
        for c in DEFAULT_CLASSES:
            cn = c["name"]
            a_c = sum(1 for s in sessions if s["course"]==cn and s["shift"]=="A")
            b_c = sum(1 for s in sessions if s["course"]==cn and s["shift"]=="B")
            c_c = sum(1 for s in sessions if s["course"]==cn and s["shift"]=="C")
            tot = a_c+b_c+c_c
            req = st.session_state.class_requirements.get(cn,0)
            row_bg = "#d4edda" if tot>=req else "#fff3cd" if tot>0 else "#f8d7da"
            tag = " ★" if c["priority"] else ""
            tbl_html += ("<tr style='background:"+row_bg+"'>"
                "<td style='padding:5px 8px'>"+cn+tag+"</td>"
                "<td style='padding:5px 8px;text-align:center'>"+(str(a_c) if a_c else "–")+"</td>"
                "<td style='padding:5px 8px;text-align:center'>"+(str(b_c) if b_c else "–")+"</td>"
                "<td style='padding:5px 8px;text-align:center'>"+(str(c_c) if c_c else "–")+"</td>"
                "<td style='padding:5px 8px;text-align:center;font-weight:bold'>"+str(tot)+" / "+str(req)+"</td></tr>")
        tbl_html += "</table>"
        st.markdown(tbl_html, unsafe_allow_html=True)
    st.markdown("---")
    if st.button("Back"): prev_step()


def calendar_view():
    st.title("Monthly Calendar View")
    sched = st.session_state.generated_schedule
    if not sched: st.warning("No schedule generated yet."); return
    month = st.session_state.schedule_month
    year = st.session_state.schedule_year
    month_dates = get_month_dates(year, month)
    sessions = sched["sessions"]
    by_date = defaultdict(list)
    for s in sessions: by_date[s["date"]].append(s)
    st.subheader(calendar.month_name[month]+" "+str(year))
    st.markdown("**Instructor Color Legend:**")
    leg_cols = st.columns(len(DEFAULT_INSTRUCTORS))
    for i, inst in enumerate(DEFAULT_INSTRUCTORS):
        col = inst_color(inst["name"])
        leg_cols[i].markdown("<div style='background:"+col["bg"]+";color:"+col["text"]+
            ";padding:4px 8px;border-radius:4px;text-align:center;font-weight:bold'>"+
            inst["name"]+"<br><small>"+DEFAULT_SHIFTS[inst["shift"]]["label"]+"</small></div>",
            unsafe_allow_html=True)
    st.markdown("---")
    weeks, week = [], []
    for _ in range(month_dates[0].weekday()): week.append(None)
    for d in month_dates:
        week.append(d)
        if len(week)==7: weeks.append(week); week=[]
    if week:
        while len(week)<7: week.append(None)
        weeks.append(week)
    hcols = st.columns(7)
    for i, h in enumerate(["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]):
        hcols[i].markdown("<div style='text-align:center;font-weight:bold;border-bottom:2px solid #ccc;padding-bottom:4px'>"+h+"</div>", unsafe_allow_html=True)
    for wk in weeks:
        cols = st.columns(7)
        for i, d in enumerate(wk):
            if d is None: cols[i].markdown("<div style='min-height:80px'></div>", unsafe_allow_html=True)
            else:
                day_sessions = sorted(by_date.get(d.isoformat(),[]), key=lambda x: x["class_start"])
                holiday = next((h for h in st.session_state.constraints["holidays"] if h["date"]==d.isoformat()), None)
                cell = "<div style='border:1px solid #ddd;border-radius:4px;padding:4px;min-height:80px;background:#fafafa'>"
                cell += "<div style='font-weight:bold;font-size:13px;margin-bottom:3px'>"+str(d.day)+"</div>"
                if holiday:
                    cell += "<div style='background:#cc0000;color:white;border-radius:3px;padding:2px 4px;font-size:10px;margin-bottom:2px'>"+holiday["label"]+"</div>"
                for s in day_sessions:
                    c = inst_color(s["instructor"])
                    cell += ("<div style='background:"+c["bg"]+";color:"+c["text"]+
                        ";border-radius:3px;padding:2px 4px;font-size:10px;margin-bottom:2px;line-height:1.3'>"
                        "<b>"+s["class_start"]+"-"+s["class_end"]+"</b><br>"+
                        s["course"]+"<br>"+s["instructor"]+" | "+s["room"]+"</div>")
                if not day_sessions and not holiday:
                    cell += "<div style='color:#aaa;font-size:10px'>No classes</div>"
                cell += "</div>"
                cols[i].markdown(cell, unsafe_allow_html=True)
        st.markdown("<div style='margin-bottom:8px'></div>", unsafe_allow_html=True)
    st.markdown("---")
    st.subheader("Export Schedule")
    if EXCEL_OK:
        excel_data = build_excel(sched, month, year)
        st.download_button("Download Excel Schedule", data=excel_data,
            file_name=calendar.month_name[month]+"_"+str(year)+"_Schedule.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("Run: pip install openpyxl to enable Excel export")


def day_detail_view():
    st.title("Day Detail View")
    sched = st.session_state.generated_schedule
    if not sched: st.warning("No schedule generated yet."); return
    month = st.session_state.schedule_month
    year = st.session_state.schedule_year
    month_dates = get_month_dates(year, month)
    sessions = sched["sessions"]
    meal_map = sched.get("meal_map", {})
    by_date = defaultdict(list)
    for s in sessions: by_date[s["date"]].append(s)
    active_dates = [d for d in month_dates if by_date.get(d.isoformat())]
    if not active_dates: st.info("No sessions scheduled."); return
    date_options = [d.strftime("%A, %B %d, %Y") for d in active_dates]
    date_lookup = {d.strftime("%A, %B %d, %Y"): d for d in active_dates}
    selected_label = st.selectbox("Select a Day", date_options)
    selected_date = date_lookup[selected_label]
    day_sessions = sorted(by_date.get(selected_date.isoformat(),[]), key=lambda x: x["class_start_min"])
    if not day_sessions: st.info("No sessions on this day."); return
    shifts_today = list(dict.fromkeys(s["shift"] for s in day_sessions))
    shift_cols = st.columns(len(shifts_today))
    for ci, sk in enumerate(shifts_today):
        sh = DEFAULT_SHIFTS[sk]
        shift_instructors = [i["name"] for i in DEFAULT_INSTRUCTORS if i["shift"]==sk]
        active_instructors = list(dict.fromkeys(s["instructor"] for s in day_sessions if s["shift"]==sk))
        sh_start, sh_end = get_shift_window(sk)
        shift_meetings = [m for m in st.session_state.constraints["meetings"]
            if m["date"]==selected_date.isoformat() and any(i in m["instructors"] for i in shift_instructors)]
        with shift_cols[ci]:
            st.markdown("### "+sh["label"])
            st.caption(sh["start"]+" - "+minutes_to_time(sh_end%(24*60)))
            for iname in active_instructors:
                inst_sessions = sorted([s for s in day_sessions if s["shift"]==sk and s["instructor"]==iname],
                    key=lambda x: x["prep_start_min"])
                col = inst_color(iname)
                meal_start = meal_map.get((iname, selected_date.isoformat()))
                st.markdown("<div style='background:"+col["bg"]+";color:"+col["text"]+
                    ";padding:3px 8px;border-radius:4px;font-weight:bold;margin-top:8px;margin-bottom:4px'>"+
                    iname+"</div>", unsafe_allow_html=True)
                events = []
                for s in inst_sessions:
                    events.append(("prep", s["prep_start_min"], s["class_start_min"], s["course"], s))
                    events.append(("class", s["class_start_min"], s["class_end_min"], s["course"], s))
                if meal_start is not None:
                    events.append(("meal", meal_start, meal_start+MEAL_MINS, "MEAL BREAK", None))
                for m in shift_meetings:
                    if iname in m["instructors"]:
                        ms = time_to_minutes(m["start"])
                        me = ms + int(m["duration_hrs"]*60)
                        events.append(("meeting", ms, me, m["label"], m))
                events.sort(key=lambda x: x[1])
                html = "<div style='font-family:monospace;font-size:12px'>"
                cursor = sh_start
                for ev in events:
                    etype, estart, eend, ename, edata = ev
                    if cursor < estart:
                        html += ("<div style='background:#e8e8e8;color:#555;padding:4px 8px;margin-bottom:2px;border-radius:3px'>" +
                            minutes_to_time(cursor)+" - "+minutes_to_time(estart)+" &nbsp; Buffer / Open</div>")
                    if etype=="prep":
                        html += ("<div style='background:#f0c040;color:#333;padding:4px 8px;margin-bottom:2px;border-radius:3px;font-weight:bold'>" +
                            minutes_to_time(estart)+" - "+minutes_to_time(eend)+" &nbsp; PREP: "+ename+"</div>")
                    elif etype=="meal":
                        html += ("<div style='background:#29b6f6;color:#000;padding:4px 8px;margin-bottom:2px;border-radius:3px;font-weight:bold'>" +
                            minutes_to_time(estart)+" - "+minutes_to_time(eend)+" &nbsp; 🍽 MEAL BREAK</div>")
                    elif etype=="meeting":
                        html += ("<div style='background:#ff9800;color:#111;padding:4px 8px;margin-bottom:2px;border-radius:3px;font-weight:bold'>" +
                            minutes_to_time(estart)+" - "+minutes_to_time(eend)+" &nbsp; 📋 "+ename+"</div>")
                    elif etype=="class":
                        c2 = inst_color(iname)
                        if edata and edata["all_day"]:
                            mid = estart + int(edata["duration_hrs"]*30)
                            html += (
                                "<div style='background:"+c2["bg"]+";color:"+c2["text"]+";padding:4px 8px;margin-bottom:2px;border-radius:3px'>" +
                                minutes_to_time(estart)+" - "+minutes_to_time(mid)+" &nbsp; "+ename+" (Part 1) | "+edata["room"]+"</div>"
                                "<div style='background:#29b6f6;color:#000;padding:4px 8px;margin-bottom:2px;border-radius:3px;font-weight:bold'>" +
                                minutes_to_time(mid)+" - "+minutes_to_time(mid+30)+" &nbsp; 🍽 MEAL BREAK (All-Day Class)</div>"
                                "<div style='background:"+c2["bg"]+";color:"+c2["text"]+";padding:4px 8px;margin-bottom:2px;border-radius:3px'>" +
                                minutes_to_time(mid+30)+" - "+minutes_to_time(eend)+" &nbsp; "+ename+" (Part 2) | "+edata["room"]+"</div>")
                        else:
                            html += ("<div style='background:"+c2["bg"]+";color:"+c2["text"]+";padding:4px 8px;margin-bottom:2px;border-radius:3px'>" +
                                minutes_to_time(estart)+" - "+minutes_to_time(eend)+" &nbsp; "+ename+" | "+(edata["room"] if edata else "")+"</div>")
                    cursor = max(cursor, eend)
                if cursor < sh_end:
                    html += ("<div style='background:#e8e8e8;color:#555;padding:4px 8px;margin-bottom:2px;border-radius:3px'>" +
                        minutes_to_time(cursor)+" - "+minutes_to_time(sh_end)+" &nbsp; Buffer / Open</div>")
                html += "</div>"
                st.markdown(html, unsafe_allow_html=True)


def build_excel(sched, month, year):
    wb = openpyxl.Workbook()
    sessions = sched["sessions"]
    meal_map = sched.get("meal_map", {})
    month_dates = get_month_dates(year, month)
    INST_HEX = {"Katie":"1a6b3c","Chris":"1a3a6b","Eric":"6b1a1a","Dave":"5a1a6b","Tempi Placeholda":"6b4e1a"}
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for sk in ["A","B","C"]:
        sh_label = DEFAULT_SHIFTS[sk]["label"]
        ws = wb.create_sheet(title=sh_label)
        shift_sessions = [s for s in sessions if s["shift"]==sk]
        shift_days_list = [d for d in month_dates if day_name(d) in DEFAULT_SHIFTS[sk]["days"]]
        if not shift_days_list: continue
        shift_instructors = [i["name"] for i in DEFAULT_INSTRUCTORS if i["shift"]==sk]
        n_inst = len(shift_instructors); n_days = len(shift_days_list)
        ws.cell(row=1,column=1,value=sh_label+" - "+calendar.month_name[month]+" "+str(year))
        ws.cell(row=1,column=1).font = Font(bold=True,size=13)
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n_days*n_inst+1)
        ws.cell(row=2,column=1,value="Time")
        ws.cell(row=2,column=1).font = Font(bold=True)
        ws.cell(row=2,column=1).fill = PatternFill("solid",fgColor="CCCCCC")
        ws.cell(row=2,column=1).border = border
        ws.column_dimensions["A"].width = 8
        for di, d in enumerate(shift_days_list):
            col_start = 2+di*n_inst; col_end = col_start+n_inst-1
            ws.cell(row=2,column=col_start,value=d.strftime("%a %m/%d"))
            ws.cell(row=2,column=col_start).font = Font(bold=True)
            ws.cell(row=2,column=col_start).fill = PatternFill("solid",fgColor="CCCCCC")
            ws.cell(row=2,column=col_start).alignment = Alignment(horizontal="center")
            ws.cell(row=2,column=col_start).border = border
            if n_inst>1: ws.merge_cells(start_row=2,start_column=col_start,end_row=2,end_column=col_end)
        for di in range(n_days):
            for ii, iname in enumerate(shift_instructors):
                col = 2+di*n_inst+ii
                c = ws.cell(row=3,column=col,value=iname)
                c.fill = PatternFill("solid",fgColor=INST_HEX.get(iname,"444444"))
                c.font = Font(color="FFFFFF",bold=True,size=8)
                c.alignment = Alignment(horizontal="center")
                c.border = border
                ws.column_dimensions[get_column_letter(col)].width = 16
        by_date_inst = defaultdict(list)
        for s in shift_sessions: by_date_inst[(s["date"],s["instructor"])].append(s)
        sh_s, sh_e = get_shift_window(sk)
        for ri, t in enumerate(range(sh_s,sh_e,30),4):
            tc = ws.cell(row=ri,column=1,value=minutes_to_time(t))
            tc.font = Font(bold=True,size=9)
            tc.fill = PatternFill("solid",fgColor="EEEEEE")
            tc.border = border
            ws.row_dimensions[ri].height = 28
            for di, d in enumerate(shift_days_list):
                for ii, iname in enumerate(shift_instructors):
                    col = 2+di*n_inst+ii
                    day_s = sorted(by_date_inst.get((d.isoformat(),iname),[]),key=lambda x:x["prep_start_min"])
                    meal_start = meal_map.get((iname,d.isoformat()))
                    cell_val,cell_fill,white_text = "","FFFFFF",False
                    meeting_here = next((m for m in st.session_state.constraints["meetings"]
                        if m["date"]==d.isoformat() and iname in m["instructors"]
                        and time_to_minutes(m["start"])<=t<time_to_minutes(m["start"])+int(m["duration_hrs"]*60)), None)
                    if meeting_here:
                        cell_val = "📋 "+meeting_here["label"]; cell_fill = "FF9800"
                    elif meal_start is not None and meal_start<=t<meal_start+MEAL_MINS:
                        cell_val = "MEAL BREAK"; cell_fill = "29B6F6"
                    else:
                        for s in day_s:
                            if s["prep_start_min"]<=t<s["class_start_min"]:
                                cell_val = "PREP\n"+s["course"]; cell_fill = "F0C040"; break
                            elif s["class_start_min"]<=t<s["class_end_min"]:
                                cell_val = s["course"]+"\n"+s["room"]
                                cell_fill = INST_HEX.get(iname,"444444"); white_text = True; break
                    cell = ws.cell(row=ri,column=col,value=cell_val)
                    cell.fill = PatternFill("solid",fgColor=cell_fill)
                    cell.font = Font(color="FFFFFF" if white_text else "000000",size=8,bold=(cell_fill in ["F0C040","FF9800"]))
                    cell.alignment = Alignment(wrap_text=True,vertical="center",horizontal="center")
                    cell.border = border
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


step = st.session_state.step
if step == 1:   step1()
elif step == 2: step2()
elif step == 3: step3()
elif step == 4: step4()
elif step == 5: step5()
elif step == 6: calendar_view()
elif step == 7: day_detail_view()
