import streamlit as st
import random
import calendar
import io
from datetime import date
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

st.set_page_config(page_title="Training Scheduler", layout="wide")

DEFAULT_CLASSES = [
    {"name": "Mech / Elec Torque",                   "duration": 4.0, "all_day": False, "room_restriction": None,                    "priority": False},
    {"name": "Safety Wire / Cable Installation",      "duration": 3.0, "all_day": False, "room_restriction": None,                    "priority": False},
    {"name": "Smart Torque",                          "duration": 2.0, "all_day": False, "room_restriction": None,                    "priority": True},
    {"name": "Threaded Insert Installation",          "duration": 3.5, "all_day": False, "room_restriction": None,                    "priority": False},
    {"name": "Fluid Fittings Installation",           "duration": 2.5, "all_day": False, "room_restriction": None,                    "priority": False},
    {"name": "Wire Harness Mate / Demate",            "duration": 3.0, "all_day": False, "room_restriction": None,                    "priority": False},
    {"name": "Wire Harness Routing and Installation", "duration": 4.0, "all_day": False, "room_restriction": None,                    "priority": False},
    {"name": "Conversion Coating",                    "duration": 1.5, "all_day": False, "room_restriction": None,                    "priority": False},
    {"name": "Application of Sealants",               "duration": 3.0, "all_day": False, "room_restriction": None,                    "priority": False},
    {"name": "Component Adhesive Bonding",            "duration": 2.0, "all_day": False, "room_restriction": None,                    "priority": False},
    {"name": "Bonding Structural",                    "duration": 4.0, "all_day": False, "room_restriction": None,                    "priority": False},
    {"name": "MPS Liquid Shim",                       "duration": 2.0, "all_day": False, "room_restriction": None,                    "priority": False},
    {"name": "Confined Space",                        "duration": 3.0, "all_day": False, "room_restriction": None,                    "priority": True},
    {"name": "Strain Gauge Installation",             "duration": 4.0, "all_day": False, "room_restriction": None,                    "priority": False},
    {"name": "Lock Out / Tag Out",                    "duration": 2.0, "all_day": False, "room_restriction": None,                    "priority": True},
    {"name": "Rynglok - Axial Swage",                 "duration": 2.0, "all_day": False, "room_restriction": None,                    "priority": False},
    {"name": "IPC 620",                               "duration": 6.0, "all_day": True,  "room_restriction": "Avionics Training Lab", "priority": False},
    {"name": "J-STD",                                 "duration": 6.0, "all_day": True,  "room_restriction": "Avionics Training Lab", "priority": False},
]

SLOT1_MIN, SLOT1_MAX = 3.0, 4.0
SLOT2_MIN, SLOT2_MAX = 1.5, 2.5

DEFAULT_SHIFTS = {
    "A1": {"label": "A1-Shift", "start": "06:30", "end": "16:00", "days": ["Monday","Tuesday","Wednesday","Thursday"]},
    "A2": {"label": "A2-Shift", "start": "06:30", "end": "16:00", "days": ["Tuesday","Wednesday","Thursday","Friday"]},
    "B":  {"label": "B-Shift",  "start": "16:30", "end": "02:00", "days": ["Monday","Tuesday","Wednesday","Thursday"]},
}

ACTIVE_SHIFTS = ["A1", "A2", "B"]

SHIFT_DAY_ORDER = {
    "A1": {"Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3},
    "A2": {"Tuesday": 0, "Wednesday": 1, "Thursday": 2, "Friday": 3},
    "B":  {"Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3},
}

DEFAULT_INSTRUCTORS = [
    {"name": "Katie",  "shift": "A1"},
    {"name": "David",  "shift": "A1"},
    {"name": "Eric",   "shift": "A1"},
    {"name": "Chris",  "shift": "A2"},
    {"name": "Erin",   "shift": "A2"},
    {"name": "Dave",   "shift": "B"},
    {"name": "Taji",   "shift": "A1", "cross_training_only": True},
]

INSTRUCTOR_COLORS = {
    "Katie":  {"bg": "#166534", "text": "#ffffff"},
    "David":  {"bg": "#1d4ed8", "text": "#ffffff"},
    "Eric":   {"bg": "#c2410c", "text": "#ffffff"},
    "Chris":  {"bg": "#0e7490", "text": "#ffffff"},
    "Erin":   {"bg": "#7c3aed", "text": "#ffffff"},
    "Dave":   {"bg": "#490b6b", "text": "#ffffff"},
    "Taji":   {"bg": "#0f766e", "text": "#ffffff"},
}
DEFAULT_COLOR    = {"bg": "#444444", "text": "#ffffff"}
DEFAULT_ROOMS    = ["Avionics Training Lab", "Galileo", "Newton", "Classroom C"]
PRIORITY_DEFAULT = 5
STANDARD_DEFAULT = 1
MEAL_MINS        = 30
PREP_MINS        = 30

NO_PREP_CLASSES = {
    "Lock Out / Tag Out", "Confined Space",
    "Wire Harness Mate / Demate", "Wire Harness Routing and Installation",
    "Rynglok - Axial Swage", "Fluid Fittings Installation",
}
SKIP_WEEKLY_FREQ = {"IPC 620", "J-STD", "Mech / Elec Torque"}

QUAL_QUALIFIED      = "Qualified"
QUAL_CROSS_TRAINING = "Cross Training"
QUAL_NOT_QUALIFIED  = "Not Qualified"
QUAL_STATES         = [QUAL_QUALIFIED, QUAL_CROSS_TRAINING, QUAL_NOT_QUALIFIED]

def init_state():
    if "step" not in st.session_state:
        st.session_state.step = 1
    if "schedule_month" not in st.session_state:
        st.session_state.schedule_month = None
    if "schedule_year" not in st.session_state:
        st.session_state.schedule_year = None
    if "class_requirements" not in st.session_state:
        st.session_state.class_requirements = {
            c["name"]: (PRIORITY_DEFAULT if c["priority"] else STANDARD_DEFAULT)
            for c in DEFAULT_CLASSES
        }
    if "constraints" not in st.session_state:
        st.session_state.constraints = {"holidays": [], "meetings": [], "pto": []}
    if "qualifications" not in st.session_state:
        all_cn = [c["name"] for c in DEFAULT_CLASSES]
        st.session_state.qualifications = {}
        for inst in DEFAULT_INSTRUCTORS:
            if inst.get("cross_training_only"):
                st.session_state.qualifications[inst["name"]] = {cn: QUAL_CROSS_TRAINING for cn in all_cn}
            else:
                st.session_state.qualifications[inst["name"]] = {cn: QUAL_QUALIFIED for cn in all_cn}
    if "generated_schedule" not in st.session_state:
        st.session_state.generated_schedule = None
    if "removed_instructors" not in st.session_state:
        st.session_state.removed_instructors = set()
    if "edited_schedule" not in st.session_state:
        st.session_state.edited_schedule = None
    if "edit_selected_idx" not in st.session_state:
        st.session_state.edit_selected_idx = None
    if "edit_panel_mode" not in st.session_state:
        st.session_state.edit_panel_mode = None
    if "edit_selected_date" not in st.session_state:
        st.session_state.edit_selected_date = None

init_state()

STEPS = ["1 - Month & Year","2 - Course Requirements","3 - Constraints","4 - Instructor Review","5 - Generate"]
st.sidebar.title("Training Scheduler")
st.sidebar.markdown("---")
for idx, label in enumerate(STEPS, 1):
    prefix = "OK " if st.session_state.step > idx else (">> " if st.session_state.step == idx else "   ")
    if st.sidebar.button(prefix + label, key="nav_"+str(idx), use_container_width=True):
        st.session_state.step = idx
if st.session_state.generated_schedule:
    st.sidebar.markdown("---")
    if st.sidebar.button("Calendar View",   key="nav_cal", use_container_width=True):
        st.session_state.step = 6
    if st.sidebar.button("Day Detail View",   key="nav_day",   use_container_width=True):
        st.session_state.step = 7
    if st.sidebar.button("Availability View", key="nav_avail", use_container_width=True):
        st.session_state.step = 8
    # Edit Schedule nav hidden until feature is complete
    # if st.sidebar.button("✏️ Edit Schedule", key="nav_edit", use_container_width=True):
    #     st.session_state.step = 9
st.sidebar.markdown("---")
st.sidebar.caption("Complete each step in order before generating.")

def next_step(): st.session_state.step = min(st.session_state.step + 1, 5)
def prev_step(): st.session_state.step = max(st.session_state.step - 1, 1)

def get_month_dates(year, month):
    _, num_days = calendar.monthrange(year, month)
    return [date(year, month, d) for d in range(1, num_days + 1)]

def get_month_weeks(year, month):
    """Return list of week-day-lists running Sun-Sat, clipped to actual month boundaries."""
    month_days = get_month_dates(year, month)
    weeks = []
    current_week = []
    for d in month_days:
        current_week.append(d)
        if d.weekday() == 5:
            weeks.append(current_week)
            current_week = []
    if current_week:
        weeks.append(current_week)
    return weeks

def day_name(d): return d.strftime("%A")

def time_to_minutes(t_str):
    h, m = map(int, t_str.split(":"))
    return h * 60 + m

def minutes_to_time(mins):
    mins = int(mins) % (24 * 60)
    return str(mins // 60).zfill(2) + ":" + str(mins % 60).zfill(2)

def inst_color(name):
    return INSTRUCTOR_COLORS.get(name, DEFAULT_COLOR)

def get_shift_window(sk):
    sh = DEFAULT_SHIFTS[sk]
    s  = time_to_minutes(sh["start"])
    e  = time_to_minutes(sh["end"])
    if e <= s: e += 24 * 60
    return s, e

def get_shift_midpoint(sk):
    s, e = get_shift_window(sk)
    return (s + e) // 2

def tuesday_meeting_type(dobj, year, month):
    if dobj.strftime("%A") != "Tuesday": return None
    tuesdays = sorted(d for d in get_month_dates(year, month) if d.strftime("%A") == "Tuesday")
    return "all_staff" if tuesdays.index(dobj) % 2 == 0 else "cop"

def get_tuesday_meeting_window(dobj, year, month):
    mtype = tuesday_meeting_type(dobj, year, month)
    if mtype is None: return None, None
    return (time_to_minutes("11:00"),
            time_to_minutes("12:30") if mtype == "all_staff" else time_to_minutes("12:00"))

def inst_blocked(iname, dobj, bstart, bend):
    ds = dobj.isoformat()
    for h in st.session_state.constraints["holidays"]:
        if h["date"] == ds: return True
    for p in st.session_state.constraints["pto"]:
        if p["instructor"] == iname and p["date"] == ds: return True
    for m in st.session_state.constraints["meetings"]:
        if m["date"] == ds and iname in m["instructors"]:
            ms = time_to_minutes(m["start"])
            me = ms + int(m["duration_hrs"] * 60)
            if not (bend <= ms or bstart >= me): return True
    return False

def is_day_blocked(iname, dobj):
    ds = dobj.isoformat()
    for h in st.session_state.constraints["holidays"]:
        if h["date"] == ds: return True
    for p in st.session_state.constraints["pto"]:
        if p["instructor"] == iname and p["date"] == ds: return True
    return False

def find_room(course, dobj, cs, ce, rsched):
    if course.get("room_restriction"):
        rooms = [course["room_restriction"]]
    else:
        rooms = [r for r in DEFAULT_ROOMS if r != "Avionics Training Lab"]
    for room in rooms:
        busy = rsched.get((room, dobj.isoformat()), [])
        if not any(not (ce <= rs or cs >= re) for rs, re, _ in busy):
            return room
    return None

def is_qualified_to_teach(iname, cn):
    return st.session_state.qualifications[iname].get(cn, QUAL_NOT_QUALIFIED) == QUAL_QUALIFIED

def is_cross_training(iname, cn):
    return st.session_state.qualifications[iname].get(cn, QUAL_NOT_QUALIFIED) == QUAL_CROSS_TRAINING

def schedule_two_class_day(sk, dobj, iname, c1_name, c2_name, cmap, rsched, isched):
    sh_s, sh_e = get_shift_window(sk)
    midpoint   = get_shift_midpoint(sk)
    c1   = cmap[c1_name]
    dur1 = int(c1["duration"] * 60)
    mtg_start, mtg_resume = get_tuesday_meeting_window(dobj, dobj.year, dobj.month)
    mtype = tuesday_meeting_type(dobj, dobj.year, dobj.month)

    def prep_for(cn):
        return 0 if cn in NO_PREP_CLASSES else PREP_MINS

    def may_end_at_meeting(cn):
        if mtype != "all_staff": return True
        return cn in ("Lock Out / Tag Out", "Confined Space")

    if c1["all_day"]:
        cs1 = sh_s; ce1 = cs1 + dur1
        if ce1 > sh_e or inst_blocked(iname, dobj, cs1, ce1): return None, None, None
        room1 = find_room(c1, dobj, cs1, ce1, rsched)
        if room1 is None: return None, None, None
        return _make_session(sk, dobj, iname, c1_name, c1, cs1, cs1, ce1, room1), None, None

    p1 = prep_for(c1_name); cs1 = sh_s + p1; ce1 = cs1 + dur1
    if mtg_start is not None:
        if cs1 < mtg_start and ce1 > mtg_start:
            if may_end_at_meeting(c1_name):
                ncs = mtg_start - dur1
                if ncs >= sh_s: cs1, ce1 = ncs, mtg_start
                else: cs1 = mtg_resume + p1; ce1 = cs1 + dur1
            else:
                cs1 = mtg_resume + p1; ce1 = cs1 + dur1
        elif mtg_start <= cs1 < mtg_resume:
            cs1 = mtg_resume + p1; ce1 = cs1 + dur1

    if ce1 > sh_e or inst_blocked(iname, dobj, cs1, ce1): return None, None, None
    room1 = find_room(c1, dobj, cs1, ce1, rsched)
    if room1 is None: return None, None, None
    s1 = _make_session(sk, dobj, iname, c1_name, c1, max(sh_s, cs1 - p1), cs1, ce1, room1)

    if c2_name is None:
        meal_start = None if (mtg_start is not None and ce1 <= mtg_start) else max(ce1, midpoint)
        return s1, None, meal_start

    c2 = cmap[c2_name]; dur2 = int(c2["duration"] * 60); p2 = prep_for(c2_name)
    if mtg_start is not None and ce1 <= mtg_start:
        ps2 = mtg_resume; cs2 = ps2 + p2; meal_start = None
    else:
        meal_start = max(ce1, midpoint); ps2 = meal_start + MEAL_MINS; cs2 = ps2 + p2
    ce2 = cs2 + dur2
    if ce2 > sh_e or inst_blocked(iname, dobj, cs2, ce2): return s1, None, meal_start
    room2 = find_room(c2, dobj, cs2, ce2, rsched)
    if room2 is None: return s1, None, meal_start
    return s1, _make_session(sk, dobj, iname, c2_name, c2, ps2, cs2, ce2, room2), meal_start


def _make_session(sk, dobj, iname, cn, course, ps, cs, ce, room, shadow_of=None):
    return {
        "date": dobj.isoformat(), "shift": sk, "course": cn,
        "instructor": iname, "room": room,
        "prep_start":  minutes_to_time(ps), "class_start": minutes_to_time(cs), "class_end": minutes_to_time(ce),
        "prep_start_min": ps, "class_start_min": cs, "class_end_min": ce,
        "duration_hrs": course["duration"], "all_day": course["all_day"],
        "shadow_of": shadow_of,
    }

def commit_session(s, rsched, isched, icount, scc):
    iname, date_iso, sk, cn = s["instructor"], s["date"], s["shift"], s["course"]
    ki = (iname, date_iso)
    isched.setdefault(ki, []).append((s["prep_start_min"], s["class_end_min"], cn))
    isched[ki].sort(key=lambda x: x[0])
    if not s.get("shadow_of"):
        rsched.setdefault((s["room"], date_iso), []).append((s["class_start_min"], s["class_end_min"], cn))
        rsched[(s["room"], date_iso)].sort(key=lambda x: x[0])
    icount[iname] = icount.get(iname, 0) + 1
    if not s.get("shadow_of"):
        scc[sk][cn] = scc[sk].get(cn, 0) + 1

def pick_slot1(pool, cmap, scc, sk, reqs, used_cn=None, day_courses=None):
    day_courses = day_courses or set()
    def shift_over(cn): return scc[sk][cn] >= max(1, reqs.get(cn, 0))
    def total_rem(cn):  return max(0, reqs.get(cn, 0) - sum(scc[s][cn] for s in ACTIVE_SHIFTS))
    def dup(cn): return 1 if cn in day_courses else 0
    candidates = [cn for cn in pool if cn != used_cn]
    t1 = sorted([cn for cn in candidates if SLOT1_MIN <= cmap[cn]["duration"] <= SLOT1_MAX and not shift_over(cn) and total_rem(cn) > 0], key=dup)
    if t1: return t1[0]
    t2 = sorted([cn for cn in candidates if SLOT1_MIN <= cmap[cn]["duration"] <= SLOT1_MAX and not shift_over(cn)], key=dup)
    if t2: return t2[0]
    t3 = sorted([cn for cn in candidates if not shift_over(cn)], key=dup)
    if t3: return t3[0]
    return sorted(candidates, key=dup)[0] if candidates else None

def pick_slot2(pool, cmap, scc, sk, reqs, exclude_cn=None, day_courses=None):
    day_courses = day_courses or set()
    def total_rem(cn): return max(0, reqs.get(cn, 0) - sum(scc[s][cn] for s in ACTIVE_SHIFTS))
    def shift_over(cn): return scc[sk][cn] >= max(1, reqs.get(cn, 0))
    def sort_key(cn): return (1 if cn in day_courses else 0, scc[sk][cn])
    pool_f = [cn for cn in pool if cn != exclude_cn]
    best = [cn for cn in pool_f if SLOT2_MIN <= cmap[cn]["duration"] <= SLOT2_MAX and total_rem(cn) > 0 and not shift_over(cn)]
    if best: best.sort(key=sort_key); return best[0]
    ok = [cn for cn in pool_f if SLOT2_MIN <= cmap[cn]["duration"] <= SLOT2_MAX and not shift_over(cn)]
    if ok: ok.sort(key=sort_key); return ok[0]
    return None

def qualified_pool_for(iname, sk, scc, reqs, cmap):
    all_cn = [c["name"] for c in DEFAULT_CLASSES]
    pool = [cn for cn in all_cn if not cmap[cn]["all_day"] and is_qualified_to_teach(iname, cn)]
    def total_rem(cn): return max(0, reqs.get(cn, 0) - sum(scc[s][cn] for s in ACTIVE_SHIFTS))
    pool.sort(key=lambda cn: (-total_rem(cn), scc[sk][cn]))
    return pool

def sort_dates_by_weekday(dates, sk):
    order = SHIFT_DAY_ORDER.get(sk, {})
    return sorted(dates, key=lambda d: order.get(day_name(d), 99))

def get_day_courses(date_iso, sessions):
    """Return the set of course names already scheduled on a given date (any shift, non-shadow)."""
    return {s["course"] for s in sessions if s["date"] == date_iso and not s.get("shadow_of")}

def generate_schedule():
    month       = st.session_state.schedule_month
    year        = st.session_state.schedule_year
    month_dates = get_month_dates(year, month)
    reqs        = dict(st.session_state.class_requirements)

    sessions = []; flags = []; rsched = {}; isched = {}; meal_map = {}
    all_cn   = [c["name"] for c in DEFAULT_CLASSES]
    cmap     = {c["name"]: c for c in DEFAULT_CLASSES}

    shift_days = {
        "A1": ["Monday","Tuesday","Wednesday","Thursday"],
        "A2": ["Tuesday","Wednesday","Thursday","Friday"],
        "B":  ["Monday","Tuesday","Wednesday","Thursday"],
    }

    removed = st.session_state.get("removed_instructors", set())
    lead_instructors = [i for i in DEFAULT_INSTRUCTORS
                        if not i.get("cross_training_only") and i["name"] not in removed]
    ct_instructors   = [i for i in DEFAULT_INSTRUCTORS
                        if i.get("cross_training_only") and i["name"] not in removed]

    ibs    = {sk: [i for i in lead_instructors if i["shift"] == sk] for sk in ACTIVE_SHIFTS}
    scc    = {s: {cn: 0 for cn in all_cn} for s in ACTIVE_SHIFTS}
    icount = {i["name"]: 0 for i in DEFAULT_INSTRUCTORS}

    sdates = {sk: sort_dates_by_weekday(
                  [d for d in month_dates if day_name(d) in shift_days[sk]], sk)
              for sk in ACTIVE_SHIFTS}

    used_slots = set()

    def try_assign_forced(iname, sk, dobj, force_c1):
        if (iname, dobj.isoformat()) in used_slots: return False
        if is_day_blocked(iname, dobj): return False
        pool = qualified_pool_for(iname, sk, scc, reqs, cmap)
        _dc  = get_day_courses(dobj.isoformat(), sessions)
        c2   = pick_slot2(pool, cmap, scc, sk, reqs, exclude_cn=force_c1, day_courses=_dc)
        s1, s2, meal_start = schedule_two_class_day(sk, dobj, iname, force_c1, c2, cmap, rsched, isched)
        if s1 is None:
            s1, s2, meal_start = schedule_two_class_day(sk, dobj, iname, force_c1, None, cmap, rsched, isched)
        if s1 is None: return False
        commit_session(s1, rsched, isched, icount, scc); sessions.append(s1)
        if s2: commit_session(s2, rsched, isched, icount, scc); sessions.append(s2)
        if meal_start is not None: meal_map[(iname, dobj.isoformat())] = meal_start
        used_slots.add((iname, dobj.isoformat()))
        return True

    def try_assign_day(iname, sk, dobj):
        if (iname, dobj.isoformat()) in used_slots: return False
        if is_day_blocked(iname, dobj): return False
        pool   = qualified_pool_for(iname, sk, scc, reqs, cmap)
        allday = [cn for cn in all_cn if cmap[cn]["all_day"] and is_qualified_to_teach(iname, cn)]
        for cn in allday:
            if max(0, reqs.get(cn, 0) - sum(scc[s][cn] for s in ACTIVE_SHIFTS)) > 0:
                s1, _, _ = schedule_two_class_day(sk, dobj, iname, cn, None, cmap, rsched, isched)
                if s1:
                    commit_session(s1, rsched, isched, icount, scc); sessions.append(s1)
                    used_slots.add((iname, dobj.isoformat())); return True
        _dc = get_day_courses(dobj.isoformat(), sessions)
        c1 = pick_slot1(pool, cmap, scc, sk, reqs, day_courses=_dc)
        if c1 is None: return False
        c2 = pick_slot2(pool, cmap, scc, sk, reqs, exclude_cn=c1, day_courses=_dc)
        s1, s2, meal_start = schedule_two_class_day(sk, dobj, iname, c1, c2, cmap, rsched, isched)
        if s1 is None:
            s1, s2, meal_start = schedule_two_class_day(sk, dobj, iname, c1, None, cmap, rsched, isched)
        if s1 is None:
            for cn in pool:
                s1, s2, meal_start = schedule_two_class_day(sk, dobj, iname, cn, None, cmap, rsched, isched)
                if s1: break
        if s1 is None: return False
        commit_session(s1, rsched, isched, icount, scc); sessions.append(s1)
        if s2: commit_session(s2, rsched, isched, icount, scc); sessions.append(s2)
        if meal_start is not None: meal_map[(iname, dobj.isoformat())] = meal_start
        used_slots.add((iname, dobj.isoformat()))
        return True

    for sk in ACTIVE_SHIFTS:
        dlist = list(sdates[sk])
        if not dlist: continue
        for course in DEFAULT_CLASSES:
            if course["all_day"]: continue
            cn = course["name"]
            if scc[sk][cn] > 0: continue
            qualified = [i for i in ibs[sk] if is_qualified_to_teach(i["name"], cn)]
            if not qualified:
                flags.append(cn + " has no qualified instructor on " + DEFAULT_SHIFTS[sk]["label"] + " - skipped.")
                continue
            placed = False
            for dobj in dlist:
                if placed: break
                for inst in sorted(qualified, key=lambda i: icount[i["name"]]):
                    if (inst["name"], dobj.isoformat()) in used_slots: continue
                    if try_assign_forced(inst["name"], sk, dobj, cn):
                        placed = True; break
            if not placed:
                flags.append("Could not schedule " + cn + " on " + DEFAULT_SHIFTS[sk]["label"] + " (minimum once).")

    for sk in ACTIVE_SHIFTS:
        dlist = list(sdates[sk])
        if not dlist: continue
        for course in DEFAULT_CLASSES:
            if not course["all_day"]: continue
            cn = course["name"]
            if scc[sk][cn] > 0: continue
            qualified = [i for i in ibs[sk] if is_qualified_to_teach(i["name"], cn)]
            if not qualified: continue
            placed = False
            for dobj in dlist:
                if placed: break
                for inst in sorted(qualified, key=lambda i: icount[i["name"]]):
                    if (inst["name"], dobj.isoformat()) in used_slots: continue
                    s1, _, _ = schedule_two_class_day(sk, dobj, inst["name"], cn, None, cmap, rsched, isched)
                    if s1:
                        commit_session(s1, rsched, isched, icount, scc); sessions.append(s1)
                        used_slots.add((inst["name"], dobj.isoformat())); placed = True; break

    n_a1 = len([i for i in lead_instructors if i["shift"] == "A1"])
    n_a2 = len([i for i in lead_instructors if i["shift"] == "A2"])
    n_b  = len([i for i in lead_instructors if i["shift"] == "B"])
    weighted_shift_order = (["A1"] * n_a1 + ["A2"] * n_a2 + ["B"] * max(n_b, 1)) * 4
    for cn in all_cn:
        req  = reqs.get(cn, 0)
        need = req - sum(scc[s][cn] for s in ACTIVE_SHIFTS)
        if need <= 0: continue
        for sk in weighted_shift_order:
            if need <= 0: break
            dlist = list(sdates[sk])
            for dobj in dlist:
                if need <= 0: break
                qualified = [i for i in ibs[sk] if is_qualified_to_teach(i["name"], cn)]
                for inst in sorted(qualified, key=lambda i: icount[i["name"]]):
                    if (inst["name"], dobj.isoformat()) in used_slots: continue
                    if try_assign_forced(inst["name"], sk, dobj, cn):
                        need -= 1; break
        if need > 0:
            flags.append("Could only schedule " + cn + " " + str(reqs.get(cn,0)-need) + "/" + str(reqs.get(cn,0)) + " times.")

    max_len = max(len(sdates[sk]) for sk in ACTIVE_SHIFTS)
    interleaved = []
    for di in range(max_len):
        for sk in ACTIVE_SHIFTS:
            if di < len(sdates[sk]):
                interleaved.append((sk, sdates[sk][di]))
    for (sk, dobj) in interleaved:
        for inst in sorted(ibs[sk], key=lambda i: icount[i["name"]]):
            try_assign_day(inst["name"], sk, dobj)

    def gap_fill_pass():
        inst_map = {i["name"]: i for i in DEFAULT_INSTRUCTORS}
        changed  = False
        for (iname, date_iso) in sorted(used_slots):
            dobj     = date.fromisoformat(date_iso)
            inst_rec = inst_map.get(iname)
            if inst_rec is None or inst_rec.get("cross_training_only"): continue
            sk = inst_rec["shift"]; _, sh_e = get_shift_window(sk)
            mtg_s, mtg_r = get_tuesday_meeting_window(dobj, dobj.year, dobj.month)
            day_blocks = isched.get((iname, date_iso), [])
            if not day_blocks: continue
            last_end   = max(b[1] for b in day_blocks)
            free_start = mtg_r if (mtg_s is not None and last_end <= mtg_s) else last_end + MEAL_MINS
            if free_start + 60 > sh_e: continue
            avail = sh_e - free_start
            def _p(cn): return 0 if cn in NO_PREP_CLASSES else PREP_MINS
            pool = [c["name"] for c in DEFAULT_CLASSES if not c["all_day"]
                    and is_qualified_to_teach(iname, c["name"])
                    and int(c["duration"] * 60) + _p(c["name"]) <= avail]
            if not pool: continue
            _gdc = get_day_courses(date_iso, sessions)
            pool.sort(key=lambda cn: (sum(scc[s][cn] for s in ACTIVE_SHIFTS), 1 if cn in _gdc else 0, scc[sk][cn]))
            for cn in pool:
                course = cmap[cn]; pv = _p(cn)
                cs = free_start + pv; ce = cs + int(course["duration"] * 60)
                if ce > sh_e or inst_blocked(iname, dobj, cs, ce): continue
                room = find_room(course, dobj, cs, ce, rsched)
                if room is None: continue
                s = _make_session(sk, dobj, iname, cn, course, free_start, cs, ce, room)
                commit_session(s, rsched, isched, icount, scc); sessions.append(s)
                if (iname, date_iso) not in meal_map and mtg_s is None:
                    meal_map[(iname, date_iso)] = last_end
                changed = True; break
        return changed
    for _ in range(3):
        if not gap_fill_pass(): break


    # ── Weekly frequency enforcement (Sun-Sat weeks, clipped to month) ──────
    month_weeks_all = get_month_weeks(year, month)
    for sk in ACTIVE_SHIFTS:
        for course in DEFAULT_CLASSES:
            cn = course["name"]
            if cn in SKIP_WEEKLY_FREQ or course["all_day"]: continue
            qualified = [i for i in ibs[sk] if is_qualified_to_teach(i["name"], cn)]
            if not qualified: continue
            for week_days in month_weeks_all:
                wdays = [d for d in week_days if day_name(d) in shift_days[sk]]
                if not wdays: continue
                week_dates_set = {d.isoformat() for d in week_days}
                already = any(
                    s["course"] == cn and s["shift"] == sk and not s.get("shadow_of")
                    and s["date"] in week_dates_set
                    for s in sessions)
                if already: continue
                placed = False
                for dobj in sorted(wdays, key=lambda d: min(icount[i["name"]] for i in qualified)):
                    if placed: break
                    for inst in sorted(qualified, key=lambda i: icount[i["name"]]):
                        if is_day_blocked(inst["name"], dobj): continue
                        if (inst["name"], dobj.isoformat()) not in used_slots:
                            if try_assign_forced(inst["name"], sk, dobj, cn):
                                placed = True; break
                        else:
                            di2 = dobj.isoformat(); _, se2 = get_shift_window(sk)
                            m2s, m2r = get_tuesday_meeting_window(dobj, dobj.year, dobj.month)
                            blks = isched.get((inst["name"], di2), [])
                            if not blks: continue
                            le   = max(b[1] for b in blks)
                            fs   = m2r if (m2s is not None and le <= m2s) else le + MEAL_MINS
                            pv   = 0 if cn in NO_PREP_CLASSES else PREP_MINS
                            ct   = fs + pv
                            ce_t = ct + int(cmap[cn]["duration"] * 60)
                            if ce_t > se2 or inst_blocked(inst["name"], dobj, ct, ce_t): continue
                            rt = find_room(cmap[cn], dobj, ct, ce_t, rsched)
                            if rt is None: continue
                            sn = _make_session(sk, dobj, inst["name"], cn, cmap[cn], fs, ct, ce_t, rt)
                            commit_session(sn, rsched, isched, icount, scc)
                            sessions.append(sn); placed = True; break
                if not placed:
                    week_label = (week_days[0].strftime("%b %d")
                                  + "–" + week_days[-1].strftime("%b %d"))
                    flags.append(
                        "Weekly freq: '" + cn + "' missing on "
                        + DEFAULT_SHIFTS[sk]["label"] + " week of " + week_label + ".")

    shadow_sessions = []
    for ct_inst in ct_instructors:
        ct_name  = ct_inst["name"]
        ct_sk    = ct_inst["shift"]
        ct_dates = sort_dates_by_weekday(
            [d for d in month_dates if day_name(d) in shift_days[ct_sk]], ct_sk)

        shadow_days_per_inst = {i["name"]: 0 for i in lead_instructors if i["shift"] == ct_sk}
        tiebreak = {i["name"]: random.random() for i in lead_instructors if i["shift"] == ct_sk}

        for dobj in ct_dates:
            if is_day_blocked(ct_name, dobj): continue
            date_iso = dobj.isoformat()

            day_leads = [s for s in sessions
                         if s["date"] == date_iso
                         and s["shift"] == ct_sk
                         and not s.get("shadow_of")
                         and is_cross_training(ct_name, s["course"])]
            if not day_leads: continue

            leads_by_inst = defaultdict(list)
            for lead_s in day_leads:
                leads_by_inst[lead_s["instructor"]].append(lead_s)

            ct_blocks = []

            def ct_overlaps(cs, ce):
                for (bs, be) in ct_blocks:
                    if not (ce <= bs or cs >= be): return True
                return False

            def course_shadow_count(cn):
                return sum(1 for x in shadow_sessions if x["course"] == cn)

            all_candidates = []
            for n in leads_by_inst:
                for s in leads_by_inst[n]:
                    all_candidates.append((n, s))
            all_candidates.sort(key=lambda x: (
                course_shadow_count(x[1]["course"]),
                shadow_days_per_inst.get(x[0], 0),
                tiebreak.get(x[0], 0)
            ))

            for lead_inst_name, lead_s in all_candidates:
                cs = lead_s["class_start_min"]
                ce = lead_s["class_end_min"]
                if inst_blocked(ct_name, dobj, cs, ce): continue
                if ct_overlaps(cs, ce): continue
                shadow = _make_session(
                    lead_s["shift"], dobj, ct_name,
                    lead_s["course"], cmap[lead_s["course"]],
                    lead_s["prep_start_min"], cs, ce,
                    lead_s["room"],
                    shadow_of=lead_s["instructor"]
                )
                ki = (ct_name, date_iso)
                isched.setdefault(ki, []).append((cs, ce, lead_s["course"]))
                isched[ki].sort(key=lambda x: x[0])
                icount[ct_name] = icount.get(ct_name, 0) + 1
                shadow_sessions.append(shadow)
                used_slots.add((ct_name, date_iso))
                ct_blocks.append((cs, ce))
                shadow_days_per_inst[lead_inst_name] = shadow_days_per_inst.get(lead_inst_name, 0) + 1
                break

    sessions.extend(shadow_sessions)
    return {"sessions": sessions, "flags": flags, "meal_map": meal_map, "isched": isched}


def build_excel(sched, month, year):
    sessions    = sched["sessions"]
    meal_map    = sched["meal_map"]
    mont_dates  = get_month_dates(year, month)
    INST_HEX    = {i["name"]: inst_color(i["name"])["bg"].replace("#","") for i in DEFAULT_INSTRUCTORS}
    thin  = Side(style="thin")
    thick = Side(style="medium")

    def mk_border(left=None, right=None, top=None, bottom=None):
        return Border(left=left or thin, right=right or thin,
                      top=top  or thin, bottom=bottom or thin)

    wb = openpyxl.Workbook()

    for sk in ACTIVE_SHIFTS:
        sh_label  = DEFAULT_SHIFTS[sk]["label"]
        sh_days   = DEFAULT_SHIFTS[sk]["days"]
        ws        = wb.create_sheet(title=sh_label)
        shift_sessions  = [s for s in sessions if s["shift"] == sk]
        shift_day_dates = sort_dates_by_weekday(
            [d for d in mont_dates if day_name(d) in sh_days], sk)
        if not shift_day_dates:
            continue
        shift_instructors = [i["name"] for i in DEFAULT_INSTRUCTORS if i["shift"] == sk]
        n_inst = len(shift_instructors)
        n_days = len(shift_day_dates)

        c = ws.cell(row=1, column=1,
                    value=sh_label + " - " + calendar.month_name[month] + " " + str(year))
        c.font      = Font(bold=True, size=13, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="1F2937")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + n_days * n_inst)
        ws.row_dimensions[1].height = 24

        tc = ws.cell(row=2, column=1, value="Time")
        tc.font      = Font(bold=True, size=10)
        tc.fill      = PatternFill("solid", fgColor="CCCCCC")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.border    = mk_border(left=thick, right=thick, top=thick, bottom=thick)
        ws.column_dimensions["A"].width = 8

        for di, d in enumerate(shift_day_dates):
            col_start = 2 + di * n_inst
            col_end   = col_start + n_inst - 1
            dc = ws.cell(row=2, column=col_start, value=d.strftime("%a  %m/%d"))
            dc.font      = Font(bold=True, size=10, color="FFFFFF")
            dc.fill      = PatternFill("solid", fgColor="2D3748")
            dc.alignment = Alignment(horizontal="center", vertical="center")
            dc.border    = mk_border(left=thick, right=thick, top=thick, bottom=thick)
            if n_inst > 1:
                ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_end)
        ws.row_dimensions[2].height = 22

        for di in range(n_days):
            for ii, iname in enumerate(shift_instructors):
                col = 2 + di * n_inst + ii
                c   = ws.cell(row=3, column=col, value=iname)
                c.fill      = PatternFill("solid", fgColor=INST_HEX.get(iname, "444444"))
                c.font      = Font(color="FFFFFF", bold=True, size=8)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = mk_border(
                    left  = thick if ii == 0          else thin,
                    right = thick if ii == n_inst - 1 else thin)
                ws.column_dimensions[get_column_letter(col)].width = 16
        ws.row_dimensions[3].height = 18

        shs, she = get_shift_window(sk)
        by_date_inst = defaultdict(list)
        for s in shift_sessions:
            by_date_inst[(s["date"], s["instructor"])].append(s)

        for ri, t in enumerate(range(shs, she, 30), 4):
            tc = ws.cell(row=ri, column=1, value=minutes_to_time(t))
            tc.font      = Font(bold=True, size=9)
            tc.fill      = PatternFill("solid", fgColor="EEEEEE")
            tc.alignment = Alignment(horizontal="center", vertical="center")
            tc.border    = mk_border(left=thick, right=thick)
            ws.row_dimensions[ri].height = 28

            for di, d in enumerate(shift_day_dates):
                for ii, iname in enumerate(shift_instructors):
                    col      = 2 + di * n_inst + ii
                    day_sess = sorted(by_date_inst.get((d.isoformat(), iname), []),
                                      key=lambda x: x["prep_start_min"])
                    meal_start = meal_map.get((iname, d.isoformat()))
                    cell_val = ""; cell_fill = "FFFFFF"; font_color = "000000"; bold = False

                    meeting_here = next(
                        (m for m in st.session_state.constraints["meetings"]
                         if m["date"] == d.isoformat() and iname in m["instructors"]
                         and time_to_minutes(m["start"]) <= t
                         < time_to_minutes(m["start"]) + int(m["duration_hrs"] * 60)), None)
                    if not meeting_here:
                        _tmt = tuesday_meeting_type(d, year, month)
                        if _tmt and time_to_minutes("11:00") <= t < time_to_minutes("12:00"):
                            meeting_here = {"label": "All Staff Meeting" if _tmt == "all_staff" else "CoP Meeting"}
                    if meeting_here:
                        cell_val = meeting_here["label"]; cell_fill = "FF9800"
                        font_color = "111111"; bold = True
                    elif meal_start is not None and meal_start <= t < meal_start + MEAL_MINS:
                        cell_val = "Meal Break"; cell_fill = "29B6F6"; bold = True
                    else:
                        matched = False
                        for s in day_sess:
                            if s["prep_start_min"] <= t < s["class_start_min"]:
                                cell_val = "PREP\n" + s["course"]
                                cell_fill = "F0C040"; font_color = "333333"; bold = True
                                matched = True; break
                            elif s["class_start_min"] <= t < s["class_end_min"]:
                                cell_val = ("(Shadow)\n" if s.get("shadow_of") else "") + s["course"] + "\n" + s["room"]
                                cell_fill = INST_HEX.get(iname, "444444")
                                font_color = "FFFFFF"; matched = True; break
                        if not matched and day_sess:
                            last_end = max(s["class_end_min"] for s in day_sess)
                            if t >= last_end:
                                cell_val = "Buffer"; cell_fill = "E0F2FE"; font_color = "555555"

                    cell = ws.cell(row=ri, column=col, value=cell_val)
                    cell.fill      = PatternFill("solid", fgColor=cell_fill)
                    cell.font      = Font(color=font_color, size=8, bold=bold)
                    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                    cell.border    = mk_border(
                        left  = thick if ii == 0          else thin,
                        right = thick if ii == n_inst - 1 else thin)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def step8():
    st.title("Availability View")
    if not st.session_state.generated_schedule:
        st.warning("No schedule generated yet."); return

    sched      = st.session_state.generated_schedule
    sessions   = sched["sessions"]
    month      = st.session_state.schedule_month
    year       = st.session_state.schedule_year
    mont_dates = get_month_dates(year, month)
    AVAIL_THRESH = 90
    TIGHT_THRESH = 60

    rows = []
    for inst in DEFAULT_INSTRUCTORS:
        if inst.get("cross_training_only"): continue
        iname    = inst["name"]
        sk       = inst["shift"]
        sh       = DEFAULT_SHIFTS[sk]
        shs, she = get_shift_window(sk)
        shift_dates = sort_dates_by_weekday(
            [d for d in mont_dates if day_name(d) in sh["days"]], sk)

        for d in shift_dates:
            date_iso = d.isoformat()
            if is_day_blocked(iname, d): continue
            inst_sessions = sorted(
                [s for s in sessions if s["instructor"] == iname
                 and s["date"] == date_iso and not s.get("shadow_of")],
                key=lambda x: x["class_start_min"])

            if not inst_sessions:
                buffer_mins  = she - shs
                status       = "Unscheduled"
                last_end_str = "--"
                h, m         = divmod(buffer_mins, 60)
                buffer_str   = f"{h}h {m:02d}m" if m else f"{h}h"
                emoji        = "🟡"
            else:
                last_end     = max(s["class_end_min"] for s in inst_sessions)
                buffer_mins  = she - last_end
                last_end_str = minutes_to_time(last_end)
                h, m         = divmod(max(buffer_mins, 0), 60)
                buffer_str   = f"{h}h {m:02d}m" if m else f"{h}h"
                if buffer_mins >= AVAIL_THRESH:
                    status = "Available"; emoji = "🟢"
                elif buffer_mins >= TIGHT_THRESH:
                    status = "Tight";     emoji = "🟠"
                else:
                    status = "Full";      emoji = "🔴"

            rows.append({"date": date_iso, "day": d.strftime("%a %m/%d"),
                         "instructor": iname, "shift": sh["label"],
                         "window": sh["start"] + " - " + sh["end"],
                         "last_end": last_end_str, "buffer_mins": buffer_mins,
                         "buffer_str": buffer_str, "status": status, "emoji": emoji})

    counts = {"Unscheduled": 0, "Available": 0, "Tight": 0, "Full": 0}
    for r in rows: counts[r["status"]] += 1
    badge_parts = [
        f"<div style='background:#1e293b;color:#fff;padding:8px 18px;border-radius:8px;font-size:0.95em'>"
        f"<b>{emoji} {label}</b>&nbsp;&nbsp;{counts[label]}</div>"
        for label, emoji in [("Unscheduled","🟡"),("Available","🟢"),("Tight","🟠"),("Full","🔴")]
    ]
    st.markdown("<div style='display:flex;gap:16px;margin-bottom:16px;flex-wrap:wrap'>"
                + "".join(badge_parts) + "</div>", unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    inst_filter   = col1.multiselect("Instructor",
                        options=sorted(set(r["instructor"] for r in rows)), default=[])
    status_filter = col2.multiselect("Status",
                        options=["Unscheduled","Available","Tight","Full"],
                        default=["Unscheduled","Available"])
    min_buf = col3.slider("Min buffer (hrs)", 0.0, 10.0, 1.5, 0.5)

    filtered = [r for r in rows
                if (not inst_filter   or r["instructor"] in inst_filter)
                and (not status_filter or r["status"]     in status_filter)
                and r["buffer_mins"] >= min_buf * 60]
    filtered.sort(key=lambda r: (r["date"], r["instructor"]))

    if not filtered:
        st.info("No results match the current filters."); return

    st.markdown(f"**{len(filtered)} entries** found")
    st.markdown("---")

    by_inst = defaultdict(list)
    for r in filtered:
        by_inst[r["instructor"]].append(r)

    for iname, inst_rows in by_inst.items():
        avail_count = sum(1 for r in inst_rows if r["status"] in ("Available","Unscheduled"))
        with st.expander(f"{iname}  -  {len(inst_rows)} days  ({avail_count} with capacity)", expanded=True):
            tbl  = "| Day | Shift Window | Last Session Ends | Buffer | Status |\n"
            tbl += "|---|---|---|---|---|\n"
            for r in inst_rows:
                tbl += f"| {r['day']} | {r['window']} | {r['last_end']} | {r['buffer_str']} | {r['emoji']} {r['status']} |\n"
            st.markdown(tbl)


def step1():
    st.title("Step 1 - Select Month & Year")
    col1, col2 = st.columns(2)
    with col1:
        year = st.selectbox("Year", list(range(2025,2030)),
                            index=list(range(2025,2030)).index(st.session_state.schedule_year or 2026))
    with col2:
        mnames    = list(calendar.month_name)[1:]
        midx      = (st.session_state.schedule_month or 3) - 1
        month_sel = st.selectbox("Month", mnames, index=midx)
        month_num = mnames.index(month_sel) + 1
    st.markdown("---")
    st.subheader("Calendar Preview")
    sdmap = {"Monday":"A1+B","Tuesday":"A1+A2+B","Wednesday":"A1+A2+B",
             "Thursday":"A1+A2+B","Friday":"A2","Saturday":"Off","Sunday":"Off"}
    dates = get_month_dates(year, month_num)
    weeks, week = [], []
    for _ in range(dates[0].weekday()): week.append(None)
    for d in dates:
        week.append(d)
        if len(week)==7: weeks.append(week); week=[]
    if week:
        while len(week)<7: week.append(None)
        weeks.append(week)
    hcols = st.columns(7)
    for i, h in enumerate(["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]):
        hcols[i].markdown("**"+h+"**")
    for wk in weeks:
        cols = st.columns(7)
        for i, d in enumerate(wk):
            if d is None: cols[i].write(" ")
            else: cols[i].markdown("**"+str(d.day)+"** "+sdmap.get(day_name(d),"Off"))
    st.markdown("---")
    if st.button("Next: Course Requirements", type="primary"):
        st.session_state.schedule_year  = year
        st.session_state.schedule_month = month_num
        next_step()

def step2():
    st.title("Step 2 - Course Requirements")
    col_left, col_right = st.columns(2)
    for i, c in enumerate(DEFAULT_CLASSES):
        default = PRIORITY_DEFAULT if c["priority"] else STANDARD_DEFAULT
        cur     = st.session_state.class_requirements.get(c["name"], default)
        tag     = " [PRIORITY]" if c["priority"] else ""
        aday    = " (All Day - Avionics Lab)" if c["all_day"] else ""
        col = col_left if i % 2 == 0 else col_right
        val = col.number_input(c["name"]+tag+" - "+str(c["duration"])+"hrs"+aday,
                               min_value=0, max_value=100, value=cur, step=1, key="req_"+c["name"])
        st.session_state.class_requirements[c["name"]] = val
    st.markdown("---")
    st.info("Total required sessions this month: "+str(sum(st.session_state.class_requirements.values())))
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Back"): prev_step()
    with col2:
        if st.button("Next: Constraints", type="primary"): next_step()

def step3():
    st.title("Step 3 - Constraints")
    inst_names  = [i["name"] for i in DEFAULT_INSTRUCTORS]
    month       = st.session_state.schedule_month or 3
    year        = st.session_state.schedule_year  or 2026
    month_dates = get_month_dates(year, month)
    date_options = [d.strftime("%A, %B %d") for d in month_dates]
    date_lookup  = {d.strftime("%A, %B %d"): d for d in month_dates}
    st.subheader("Holidays / Blackout Days")
    with st.form("hform", clear_on_submit=True):
        hc1, hc2 = st.columns([3,1])
        hdate  = hc1.selectbox("Date", date_options, key="hdate")
        hlabel = hc1.text_input("Label", key="hlabel")
        hc2.markdown("<br><br>", unsafe_allow_html=True)
        if hc2.form_submit_button("Add") and hlabel:
            st.session_state.constraints["holidays"].append({"date": date_lookup[hdate].isoformat(), "label": hlabel})
    for i, h in enumerate(st.session_state.constraints["holidays"]):
        c1, c2 = st.columns([5,1])
        c1.markdown("Holiday: **"+h["label"]+"** on "+h["date"])
        if c2.button("Remove", key="dh_"+str(i)):
            st.session_state.constraints["holidays"].pop(i); st.rerun()
    st.markdown("---")
    st.subheader("Instructor PTO")
    with st.form("pform", clear_on_submit=True):
        pc1, pc2, pc3 = st.columns(3)
        pto_inst = pc1.selectbox("Instructor", inst_names, key="ptoi")
        pto_date = pc2.selectbox("Date", date_options, key="ptod")
        pc3.markdown("<br><br>", unsafe_allow_html=True)
        if pc3.form_submit_button("Add PTO"):
            st.session_state.constraints["pto"].append({"instructor": pto_inst, "date": date_lookup[pto_date].isoformat()})
    for i, p in enumerate(st.session_state.constraints["pto"]):
        c1, c2 = st.columns([5,1])
        c1.markdown("PTO: **"+p["instructor"]+"** on "+p["date"])
        if c2.button("Remove", key="dp_"+str(i)):
            st.session_state.constraints["pto"].pop(i); st.rerun()
    st.markdown("---")
    st.subheader("Meetings")
    with st.form("mform", clear_on_submit=True):
        mc1, mc2 = st.columns(2)
        mdate  = mc1.selectbox("Date", date_options, key="mdate")
        minst  = mc1.multiselect("Instructors", inst_names, key="minst")
        mlabel = mc1.text_input("Label", value="Team Meeting", key="mlabel")
        mstart = mc2.time_input("Start Time", key="mstart")
        mdur   = mc2.number_input("Duration (hrs)", min_value=0.5, max_value=8.0, value=1.0, step=0.5, key="mdur")
        if st.form_submit_button("Add Meeting") and minst and mstart:
            st.session_state.constraints["meetings"].append({
                "date": date_lookup[mdate].isoformat(), "instructors": minst, "label": mlabel,
                "start": mstart.strftime("%H:%M"), "duration_hrs": float(mdur)})
    for i, m in enumerate(st.session_state.constraints["meetings"]):
        c1, c2 = st.columns([5,1])
        c1.markdown("Meeting: **"+m["label"]+"** on "+m["date"]+" at "+m["start"]+" ("+str(m["duration_hrs"])+"hr) - "+", ".join(m["instructors"]))
        if c2.button("Remove", key="dm_"+str(i)):
            st.session_state.constraints["meetings"].pop(i); st.rerun()
    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Back"): prev_step()
    with col2:
        if st.button("Next: Instructor Review", type="primary"): next_step()

def step4():
    st.title("Step 4 - Instructor Review & Qualifications")
    all_cn = [c["name"] for c in DEFAULT_CLASSES]
    for inst in DEFAULT_INSTRUCTORS:
        name       = inst["name"]
        sh         = DEFAULT_SHIFTS[inst["shift"]]
        is_removed = name in st.session_state.removed_instructors
        _lbl = ("🚫 " if is_removed else "") + name + " - " + sh["label"] + " (" + sh["start"] + " to " + sh["end"] + ", " + ", ".join(sh["days"]) + ")" + (" [REMOVED FROM ROTATION]" if is_removed else "")
        with st.expander(_lbl, expanded=False):
            _removed_cb = st.checkbox(
                "🚫 Remove from rotation (exclude entirely from scheduling)",
                value=is_removed, key="remove_"+name)
            if _removed_cb:
                st.session_state.removed_instructors.add(name)
                st.warning("⚠️ " + name + " will be excluded from the generated schedule.")
            else:
                st.session_state.removed_instructors.discard(name)
            st.markdown("---")
            cols = st.columns(3)
            for j, cn in enumerate(all_cn):
                cur = st.session_state.qualifications[name].get(cn, QUAL_NOT_QUALIFIED)
                nv  = cols[j%3].selectbox(cn, QUAL_STATES, index=QUAL_STATES.index(cur), key="q_"+name+"_"+cn)
                st.session_state.qualifications[name][cn] = nv
    st.markdown("---")
    tbl = "| Instructor | Shift | Qualified | Cross Training | Not Qualified |\n|---|---|---|---|---|\n"
    for inst in DEFAULT_INSTRUCTORS:
        name  = inst["name"]
        qvals = list(st.session_state.qualifications[name].values())
        n_q   = sum(1 for v in qvals if v == QUAL_QUALIFIED)
        n_ct  = sum(1 for v in qvals if v == QUAL_CROSS_TRAINING)
        n_nq  = sum(1 for v in qvals if v == QUAL_NOT_QUALIFIED)
        tbl  += "| "+name+" | "+DEFAULT_SHIFTS[inst["shift"]]["label"]+" | "+str(n_q)+" | "+str(n_ct)+" | "+str(n_nq)+" |\n"
    st.markdown(tbl)
    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Back"): prev_step()
    with col2:
        if st.button("Next: Generate", type="primary"): next_step()

def step5():
    st.title("Step 5 - Review & Generate")
    month = st.session_state.schedule_month
    year  = st.session_state.schedule_year
    if not month or not year:
        st.warning("Please go back to Step 1 and select a month and year.")
        if st.button("Back to Step 1"): st.session_state.step = 1
        return
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Schedule Period")
        st.markdown("**"+calendar.month_name[month]+" "+str(year)+"**")
        st.subheader("Course Requirements")
        for c in DEFAULT_CLASSES:
            count = st.session_state.class_requirements.get(c["name"], STANDARD_DEFAULT)
            tag   = " [P]" if c["priority"] else ""
            st.markdown("- "+c["name"]+tag+": **"+str(count)+"**")
    with col2:
        st.subheader("Constraints")
        h = st.session_state.constraints["holidays"]
        p = st.session_state.constraints["pto"]
        m = st.session_state.constraints["meetings"]
        st.markdown("Holidays: "+str(len(h))+" | PTO: "+str(len(p))+" | Meetings: "+str(len(m)))
        st.subheader("Instructor Qualifications")
        for inst in DEFAULT_INSTRUCTORS:
            name  = inst["name"]
            qvals = list(st.session_state.qualifications[name].values())
            n_q   = sum(1 for v in qvals if v == QUAL_QUALIFIED)
            n_ct  = sum(1 for v in qvals if v == QUAL_CROSS_TRAINING)
            st.markdown("- "+name+" ("+DEFAULT_SHIFTS[inst["shift"]]["label"]+"): "+str(n_q)+" Qualified, "+str(n_ct)+" Cross Training")
    st.markdown("---")
    st.info("Total minimum required sessions: "+str(sum(st.session_state.class_requirements.values())))
    if st.button("GENERATE SCHEDULE", type="primary", use_container_width=True):
        with st.spinner("Building schedule..."):
            result = generate_schedule()
        st.session_state.generated_schedule = result
        st.success("Schedule generated! Use the sidebar to view Calendar or Day Detail.")
        st.rerun()
    if st.session_state.generated_schedule:
        sched    = st.session_state.generated_schedule
        sessions = sched["sessions"]
        if sched["flags"]:
            _nf = len(sched["flags"])
            with st.expander(f"⚠️ {_nf} Scheduling Flag{'s' if _nf != 1 else ''} — click to expand", expanded=False):
                for f in sched["flags"]: st.warning(f)
        st.markdown("---")
        st.subheader("Quick Summary")
        st.markdown("#### Instructor Load")
        tbl = "| Instructor | Shift | Lead Sessions | Shadow Sessions | Hours |\n|--|--|--|--|--|\n"
        for inst in DEFAULT_INSTRUCTORS:
            name     = inst["name"]
            inst_s   = [s for s in sessions if s["instructor"]==name and not s.get("shadow_of")]
            shadow_s = [s for s in sessions if s["instructor"]==name and s.get("shadow_of")]
            total_hrs = sum(s["duration_hrs"] for s in inst_s)
            ct_tag    = " (CT)" if inst.get("cross_training_only") else ""
            tbl += "| "+name+ct_tag+" | "+DEFAULT_SHIFTS[inst["shift"]]["label"]+" | "+str(len(inst_s))+" | "+str(len(shadow_s))+" | "+str(round(total_hrs,1))+" hrs |\n"
        st.markdown(tbl)

        # ── Taji Shadow Breakdowns ─────────────────────────────────────────
        taji_shadows = [s for s in sessions if s["instructor"]=="Taji" and s.get("shadow_of")]
        if taji_shadows:
            # Instructor breakdown (inline caption)
            shadow_by_inst      = defaultdict(int)
            shadow_days_by_inst = defaultdict(set)
            for s in taji_shadows:
                shadow_by_inst[s["shadow_of"]] += 1
                shadow_days_by_inst[s["shadow_of"]].add(s["date"])
            breakdown = " | ".join(
                n + ": " + str(shadow_by_inst[n]) + " sessions / " + str(len(shadow_days_by_inst[n])) + " days"
                for n in sorted(shadow_by_inst.keys())
            )
            st.caption("Taji shadow breakdown → " + breakdown)

            # Class breakdown (collapsible table, sorted by most-shadowed first)
            class_shadow_count = defaultdict(int)
            for s in taji_shadows:
                class_shadow_count[s["course"]] += 1
            sorted_classes = sorted(class_shadow_count.items(), key=lambda x: -x[1])
            tbl_class = "| Course | Times Shadowed |\n|---|---|\n"
            for cn, cnt in sorted_classes:
                tbl_class += "| " + cn + " | " + str(cnt) + " |\n"
            with st.expander("Taji Shadow Class Breakdown"):
                st.markdown(tbl_class)

        st.markdown("#### Course Coverage")
        hdr  = "| Course | "+" | ".join(DEFAULT_SHIFTS[sk]["label"] for sk in ACTIVE_SHIFTS)+" | Total / Req |\n"
        sep  = "|--|"+ "|".join(["--"]*len(ACTIVE_SHIFTS)) +"|--|\n"
        rows = []
        for c in DEFAULT_CLASSES:
            cn  = c["name"]
            req = st.session_state.class_requirements.get(cn, 0)
            tag = " [P]" if c["priority"] else ""
            counts = []; tot = 0
            for sk in ACTIVE_SHIFTS:
                n = sum(1 for s in sessions if s["course"]==cn and s["shift"]==sk and not s.get("shadow_of"))
                counts.append(str(n) if n else "-"); tot += n
            rows.append("| "+cn+tag+" | "+" | ".join(counts)+" | "+str(tot)+" / "+str(req)+" |")
        st.markdown(hdr+sep+"\n".join(rows))

        st.markdown("#### Weekly Frequency Check")
        st.caption("\u2705 = taught that week on that shift  \u274c = missing  \u2014 = exempt (IPC 620, J-STD, Mech/Elec Torque)")
        _month_weeks = get_month_weeks(year, month)
        _wh = []
        for _wk_days in _month_weeks:
            _ws = _wk_days[0]; _we = _wk_days[-1]
            _partial = " *(partial)*" if len(_wk_days) < 7 else ""
            _wh.append(_ws.strftime("%b %d") + "\u2013" + _we.strftime("%b %d") + _partial)
        _wf_hdr = "| Class | " + " | ".join(_wh) + " |\n"
        _wf_sep = "|--|" + "|".join(["--"] * len(_month_weeks)) + "|\n"
        _wf_rows = []
        for _c in DEFAULT_CLASSES:
            _cn = _c["name"]
            if _cn in SKIP_WEEKLY_FREQ or _c["all_day"]:
                _wf_rows.append("| " + _cn + " | " + " | ".join(["\u2014"] * len(_month_weeks)) + " |")
                continue
            _cells = []
            for _wk_days in _month_weeks:
                _week_dates_set = {_d.isoformat() for _d in _wk_days}
                _hits = set()
                for _s in sessions:
                    if _s["course"] == _cn and not _s.get("shadow_of"):
                        if _s["date"] in _week_dates_set:
                            _hits.add(_s["shift"])
                _cells.append("\u2705 " + ", ".join(sorted(_hits)) if _hits else "\u274c")
            _wf_rows.append("| " + _cn + " | " + " | ".join(_cells) + " |")
        st.markdown(_wf_hdr + _wf_sep + "\n".join(_wf_rows))

def step6():
    st.title("Monthly Calendar View")
    if not st.session_state.generated_schedule:
        st.warning("No schedule generated yet."); return
    sched    = st.session_state.generated_schedule
    sessions = sched["sessions"]
    month    = st.session_state.schedule_month
    year     = st.session_state.schedule_year
    st.markdown("### "+calendar.month_name[month]+" "+str(year))

    # ── Calendar Filters ──────────────────────────────────────────────
    all_instructors = [i["name"] for i in DEFAULT_INSTRUCTORS
                       if not i.get("cross_training_only")]
    all_courses = sorted(set(s["course"] for s in sessions
                            if not s.get("shadow_of")))
    fc1, fc2 = st.columns(2)
    filter_insts   = fc1.multiselect("Filter by Instructor",
                         options=all_instructors, default=[],
                         placeholder="All instructors shown")
    filter_courses = fc2.multiselect("Filter by Course",
                         options=all_courses, default=[],
                         placeholder="All courses shown")
    def is_highlighted(s):
        if not filter_insts and not filter_courses: return True
        inst_match   = (not filter_insts   or s["instructor"] in filter_insts)
        course_match = (not filter_courses or s["course"]     in filter_courses)
        return inst_match and course_match
    st.markdown("**Instructor Color Legend:**")
    legend_html = "<div style='display:flex;flex-wrap:wrap;gap:8px;margin-bottom:16px'>"
    for inst in DEFAULT_INSTRUCTORS:
        col = inst_color(inst["name"])
        ct  = " (CT)" if inst.get("cross_training_only") else ""
        legend_html += ("<div style='background:"+col["bg"]+";color:"+col["text"]+";"
                        "padding:8px 16px;border-radius:6px;text-align:center;min-width:120px'>"
                        "<b>"+inst["name"]+ct+"</b><br><small>"+DEFAULT_SHIFTS[inst["shift"]]["label"]+"</small></div>")
    legend_html += "</div>"
    st.markdown(legend_html, unsafe_allow_html=True)
    st.markdown("---")
    dates = get_month_dates(year, month)
    sess_by_date = defaultdict(list)
    for s in sessions: sess_by_date[s["date"]].append(s)
    weeks, week = [], []
    for _ in range(dates[0].weekday()): week.append(None)
    for d in dates:
        week.append(d)
        if len(week)==7: weeks.append(week); week=[]
    if week:
        while len(week)<7: week.append(None)
        weeks.append(week)
    hcols = st.columns(7)
    for i, h in enumerate(["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]):
        hcols[i].markdown("<div style='text-align:center;font-weight:bold;padding:4px'>"+h+"</div>", unsafe_allow_html=True)
    for wk in weeks:
        cols = st.columns(7)
        for i, d in enumerate(wk):
            if d is None:
                cols[i].markdown("<div style='min-height:80px'></div>", unsafe_allow_html=True)
            else:
                day_sessions = sorted(sess_by_date.get(d.isoformat(), []), key=lambda x: x["class_start_min"])
                cell = "<div style='border:1px solid #ddd;border-radius:4px;padding:4px;min-height:80px'>"
                cell += "<div style='font-weight:bold;margin-bottom:4px'>"+str(d.day)+"</div>"
                for s in day_sessions:
                    col         = inst_color(s["instructor"])
                    shadow_tag  = " [Shadow]" if s.get("shadow_of") else ""
                    highlighted = is_highlighted(s)
                    bg  = col["bg"]    if highlighted else "#e5e7eb"
                    fg  = col["text"]  if highlighted else "#9ca3af"
                    op  = "1.0"        if highlighted else "0.45"
                    cell += (f"<div style='background:{bg};color:{fg};opacity:{op};"
                             f"border-radius:3px;padding:2px 4px;margin-bottom:2px;font-size:0.72em'>"
                             f"<b>{s['class_start']} - {s['class_end']}</b><br>"
                             f"{s['course']}<br>"
                             f"{s['instructor']}{shadow_tag} | {s['room']}</div>")
                # Availability badges: green +Xh for instructors with 90+ min buffer
                shown_insts = list(dict.fromkeys(
                    s["instructor"] for s in day_sessions if not s.get("shadow_of")))
                badge_line = ""
                for iname in shown_insts:
                    inst_rec = next((x for x in DEFAULT_INSTRUCTORS if x["name"] == iname), None)
                    if not inst_rec or inst_rec.get("cross_training_only"): continue
                    _, she_i = get_shift_window(inst_rec["shift"])
                    inst_day = [s for s in day_sessions
                                if s["instructor"] == iname and not s.get("shadow_of")]
                    if inst_day:
                        last_e = max(s["class_end_min"] for s in inst_day)
                        buf_m  = she_i - last_e
                        if buf_m >= 90:
                            bh, bm = divmod(buf_m, 60)
                            blabel = f"+{bh}h{bm:02d}" if bm else f"+{bh}h"
                            badge_line += (f"<span style='font-size:0.62em;background:#dcfce7;"
                                           f"color:#166534;border-radius:3px;padding:1px 4px;"
                                           f"margin-right:2px'>{iname} {blabel}</span>")
                if badge_line:
                    cell += f"<div style='margin-top:3px'>{badge_line}</div>"
                cell += "</div>"
                cols[i].markdown(cell, unsafe_allow_html=True)
    if EXCEL_OK:
        st.markdown("---")
        if st.button("Download Excel"):
            buf = io.BytesIO()
            wb  = build_excel(sched, month, year)
            wb.save(buf); buf.seek(0)
            st.download_button(
                "Download Excel Schedule", buf,
                file_name=calendar.month_name[month]+str(year)+"_Schedule.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

def step7():
    st.title("Day Detail View")
    if not st.session_state.generated_schedule:
        st.warning("No schedule generated yet."); return
    sched    = st.session_state.generated_schedule
    sessions = sched["sessions"]
    meal_map = sched["meal_map"]
    month    = st.session_state.schedule_month
    year     = st.session_state.schedule_year
    dates        = get_month_dates(year, month)
    active_dates = sorted({s["date"] for s in sessions})
    date_options = [d.isoformat() for d in dates if d.isoformat() in active_dates]
    if not date_options: st.info("No sessions scheduled."); return
    sel_date = st.selectbox("Select a Day", date_options,
                            format_func=lambda d: date.fromisoformat(d).strftime("%A, %B %d, %Y"))
    day_sessions = sorted([s for s in sessions if s["date"] == sel_date], key=lambda s: s["class_start_min"])
    if not day_sessions: st.info("No sessions on this day."); return
    meetings_today = [m for m in st.session_state.constraints["meetings"] if m["date"] == sel_date]

    def render_instructor_timeline(iname, sk, all_sessions):
        inst_sessions = sorted([s for s in all_sessions if s["instructor"] == iname], key=lambda s: s["class_start_min"])
        if not inst_sessions: return
        col = inst_color(iname)
        sh  = DEFAULT_SHIFTS[sk]
        sh_s, sh_e = get_shift_window(sk)
        st.markdown("<div style='background:"+col["bg"]+";color:"+col["text"]+";padding:6px 12px;border-radius:4px;font-weight:bold;margin-top:8px'>"+iname+"</div>", unsafe_allow_html=True)
        st.markdown("<div style='color:#888;font-size:0.82em;margin:2px 0 6px 4px'>"+sh["start"]+" - "+sh["end"]+"</div>", unsafe_allow_html=True)
        meal_min      = meal_map.get((iname, sel_date))
        inst_meetings = [m for m in meetings_today if iname in m["instructors"]]
        timeline = []
        for s in inst_sessions:
            if s["prep_start_min"] != s["class_start_min"]:
                timeline.append(("prep", s["prep_start_min"], s["class_start_min"], s["course"]))
            shadow_label = " (Shadowing "+s["shadow_of"]+")" if s.get("shadow_of") else ""
            shadows_on   = [x["instructor"] for x in inst_sessions
                            if x.get("shadow_of") == iname and x["course"] == s["course"]
                            and x["class_start_min"] == s["class_start_min"]]
            shadow_with  = (" (w/ Shadow: "+", ".join(shadows_on)+")") if shadows_on else ""
            timeline.append(("class", s["class_start_min"], s["class_end_min"],
                             s["course"]+" | "+s["room"]+shadow_label+shadow_with))
            if meal_min is not None and s["class_end_min"] == meal_min:
                timeline.append(("meal", meal_min, meal_min + MEAL_MINS, ""))
        for m in inst_meetings:
            ms = time_to_minutes(m["start"]); me = ms + int(m["duration_hrs"] * 60)
            timeline.append(("meeting", ms, me, m["label"]))
        timeline.sort(key=lambda x: x[1])
        final = []; prev_end = sh_s
        for event in timeline:
            if event[1] > prev_end + 1:
                final.append(("buffer", prev_end, event[1], ""))
            final.append(event); prev_end = event[2]
        if prev_end < sh_e:
            final.append(("buffer", prev_end, sh_e, ""))
        ROW_STYLES = {
            "prep":    "background:#d97706;color:#fff;padding:3px 10px;border-radius:3px;margin-bottom:2px;font-size:0.82em",
            "class":   "background:"+col["bg"]+";color:"+col["text"]+";padding:5px 10px;border-radius:3px;margin-bottom:2px;font-size:0.85em",
            "meal":    "background:#0ea5e9;color:#fff;padding:3px 10px;border-radius:3px;margin-bottom:2px;font-size:0.82em",
            "meeting": "background:#f59e0b;color:#000;padding:3px 10px;border-radius:3px;margin-bottom:2px;font-size:0.82em",
            "buffer":  "background:#e5e7eb;color:#555;padding:3px 10px;border-radius:3px;margin-bottom:2px;font-size:0.82em",
        }
        for (etype, es, ee, label) in final:
            time_str = minutes_to_time(es)+" - "+minutes_to_time(ee)
            style    = ROW_STYLES[etype]
            if etype == "meal":     content = time_str+"&nbsp;&nbsp;&nbsp;MEAL BREAK"
            elif etype == "buffer": content = time_str+"&nbsp;&nbsp;&nbsp;Buffer / Open"
            elif etype == "meeting": content = time_str+"&nbsp;&nbsp;&nbsp;"+label
            elif etype == "prep":   content = time_str+"&nbsp;&nbsp;&nbsp;PREP: "+label
            else:                   content = time_str+"&nbsp;&nbsp;&nbsp;"+label
            st.markdown("<div style='"+style+"'>"+content+"</div>", unsafe_allow_html=True)

    left_col, right_col = st.columns(2)
    with left_col:
        for sk in ["A1", "A2"]:
            sk_sessions = [s for s in day_sessions if s["shift"] == sk]
            if not sk_sessions: continue
            st.markdown("## "+DEFAULT_SHIFTS[sk]["label"])
            st.markdown("<div style='color:#888;font-size:0.85em'>"+DEFAULT_SHIFTS[sk]["start"]+" - "+DEFAULT_SHIFTS[sk]["end"]+"</div>", unsafe_allow_html=True)
            for inst in DEFAULT_INSTRUCTORS:
                if inst["shift"] == sk:
                    render_instructor_timeline(inst["name"], sk, day_sessions)
    with right_col:
        sk_sessions = [s for s in day_sessions if s["shift"] == "B"]
        if sk_sessions:
            st.markdown("## "+DEFAULT_SHIFTS["B"]["label"])
            st.markdown("<div style='color:#888;font-size:0.85em'>"+DEFAULT_SHIFTS["B"]["start"]+" - "+DEFAULT_SHIFTS["B"]["end"]+"</div>", unsafe_allow_html=True)
            for inst in DEFAULT_INSTRUCTORS:
                if inst["shift"] == "B":
                    render_instructor_timeline(inst["name"], "B", day_sessions)


# =======================================================================
#  STEP 9 - EDIT SCHEDULE
# =======================================================================

def _init_edited():
    import copy
    if st.session_state.edited_schedule is None and st.session_state.generated_schedule:
        st.session_state.edited_schedule = copy.deepcopy(st.session_state.generated_schedule)


def _detect_conflicts(sessions):
    conflicts = set()
    for i, a in enumerate(sessions):
        for j, b in enumerate(sessions):
            if i >= j: continue
            if a["date"] != b["date"]: continue
            overlap = not (a["class_end_min"] <= b["class_start_min"]
                           or a["class_start_min"] >= b["class_end_min"])
            if overlap:
                if a["instructor"] == b["instructor"]:
                    conflicts.add(i); conflicts.add(j)
                if (a["room"] == b["room"]
                        and not a.get("shadow_of") and not b.get("shadow_of")):
                    conflicts.add(i); conflicts.add(j)
    return conflicts


@st.dialog("Edit Session", width="large")
def _edit_dialog(sidx, sessions, month, year):
    cmap2  = {c["name"]: c for c in DEFAULT_CLASSES}
    s      = sessions[sidx]
    st.markdown("**Editing:** " + s["course"] + " — " + s["instructor"]
                + " on " + date.fromisoformat(s["date"]).strftime("%A, %B %d"))
    st.markdown("---")
    all_cn2    = [c["name"] for c in DEFAULT_CLASSES]
    new_course = st.selectbox("Course", all_cn2,
        index=all_cn2.index(s["course"]) if s["course"] in all_cn2 else 0)
    sk         = s["shift"]
    sh_insts   = [i["name"] for i in DEFAULT_INSTRUCTORS if i["shift"] == sk]
    new_inst   = st.selectbox("Instructor", sh_insts,
        index=sh_insts.index(s["instructor"]) if s["instructor"] in sh_insts else 0)
    new_room   = st.selectbox("Room", DEFAULT_ROOMS,
        index=DEFAULT_ROOMS.index(s["room"]) if s["room"] in DEFAULT_ROOMS else 0)
    valid_dates = [d for d in get_month_dates(year, month)
                   if day_name(d) in DEFAULT_SHIFTS[sk]["days"]]
    date_strs   = [d.isoformat() for d in valid_dates]
    new_date    = st.selectbox("Date", date_strs,
        index=date_strs.index(s["date"]) if s["date"] in date_strs else 0,
        format_func=lambda x: date.fromisoformat(x).strftime("%A, %B %d"))
    sh_s2, sh_e2 = get_shift_window(sk)
    time_opts    = list(range(sh_s2, sh_e2, 30))
    time_strs2   = [minutes_to_time(t) for t in time_opts]
    cur_ti       = time_opts.index(s["class_start_min"]) if s["class_start_min"] in time_opts else 0
    new_st_str   = st.selectbox("Start Time", time_strs2, index=cur_ti)
    new_st_min   = time_to_minutes(new_st_str)
    dur_min      = int(cmap2[new_course]["duration"] * 60)
    new_end_min  = new_st_min + dur_min
    st.caption("Ends at " + minutes_to_time(new_end_min)
               + "  (" + str(cmap2[new_course]["duration"]) + " hrs)")
    st.markdown("---")
    ca, cb, cc = st.columns([2, 1, 1])
    if ca.button("💾 Save", type="primary", use_container_width=True):
        sessions[sidx].update({
            "course": new_course, "instructor": new_inst,
            "room": new_room, "date": new_date,
            "class_start": new_st_str, "class_start_min": new_st_min,
            "class_end": minutes_to_time(new_end_min), "class_end_min": new_end_min,
            "prep_start_min": max(sh_s2, new_st_min - PREP_MINS),
            "prep_start": minutes_to_time(max(sh_s2, new_st_min - PREP_MINS)),
            "duration_hrs": cmap2[new_course]["duration"],
        })
        st.session_state.edit_selected_idx = None
        st.session_state.edit_panel_mode   = None
        st.query_params.clear(); st.rerun()
    if cb.button("🗑️ Remove", use_container_width=True):
        sessions.pop(sidx)
        st.session_state.edit_selected_idx = None
        st.session_state.edit_panel_mode   = None
        st.query_params.clear(); st.rerun()
    if cc.button("✖ Cancel", use_container_width=True):
        st.session_state.edit_selected_idx = None
        st.session_state.edit_panel_mode   = None
        st.query_params.clear(); st.rerun()


@st.dialog("Add New Class", width="large")
def _add_dialog(sessions, month, year, default_date=None):
    cmap2     = {c["name"]: c for c in DEFAULT_CLASSES}
    all_cn2   = [c["name"] for c in DEFAULT_CLASSES]
    new_course = st.selectbox("Course", all_cn2)
    new_sk     = st.selectbox("Shift", ACTIVE_SHIFTS,
                              format_func=lambda x: DEFAULT_SHIFTS[x]["label"])
    sh_insts   = [i["name"] for i in DEFAULT_INSTRUCTORS if i["shift"] == new_sk]
    new_inst   = st.selectbox("Instructor", sh_insts)
    new_room   = st.selectbox("Room", DEFAULT_ROOMS)
    valid_d    = [d for d in get_month_dates(year, month)
                  if day_name(d) in DEFAULT_SHIFTS[new_sk]["days"]]
    date_strs  = [d.isoformat() for d in valid_d]
    def_idx    = date_strs.index(default_date) if default_date and default_date in date_strs else 0
    new_date   = st.selectbox("Date", date_strs, index=def_idx,
        format_func=lambda x: date.fromisoformat(x).strftime("%A, %B %d"))
    sh_s3, sh_e3 = get_shift_window(new_sk)
    time_opts    = list(range(sh_s3, sh_e3, 30))
    time_strs3   = [minutes_to_time(t) for t in time_opts]
    new_st_str   = st.selectbox("Start Time", time_strs3)
    new_st_min   = time_to_minutes(new_st_str)
    dur_min      = int(cmap2[new_course]["duration"] * 60)
    new_end      = new_st_min + dur_min
    st.caption("Ends at " + minutes_to_time(new_end)
               + "  (" + str(cmap2[new_course]["duration"]) + " hrs)")
    st.markdown("---")
    ca2, cb2 = st.columns(2)
    if ca2.button("✅ Add", type="primary", use_container_width=True):
        sessions.append(_make_session(
            new_sk, date.fromisoformat(new_date), new_inst,
            new_course, cmap2[new_course],
            max(sh_s3, new_st_min - PREP_MINS), new_st_min, new_end, new_room))
        st.session_state.edit_panel_mode = None
        st.rerun()
    if cb2.button("✖ Cancel", use_container_width=True):
        st.session_state.edit_panel_mode = None
        st.rerun()


def _build_timeline_html(sel_date_iso, sessions, conflicts):
    ROW_PX = 40; SLOT_MINS = 30; TIME_COL = 64; INST_COL = 160
    sel_date   = date.fromisoformat(sel_date_iso)
    dname      = day_name(sel_date)
    act_shifts = [sk for sk in ACTIVE_SHIFTS if dname in DEFAULT_SHIFTS[sk]["days"]]
    if not act_shifts:
        return "<p>No shifts scheduled on this day.</p>", 100
    day_sess  = [s for s in sessions if s["date"] == sel_date_iso]
    sess_idx  = {id(s): i for i, s in enumerate(sessions)}
    t_start   = min(get_shift_window(sk)[0] for sk in act_shifts)
    t_end     = max(get_shift_window(sk)[1] for sk in act_shifts)
    slots     = list(range(t_start, t_end, SLOT_MINS))
    inst_list = [(i["name"], i["shift"])
                 for sk in act_shifts
                 for i in DEFAULT_INSTRUCTORS if i["shift"] == sk]
    total_w   = TIME_COL + len(inst_list) * INST_COL
    total_h   = 70 + len(slots) * (ROW_PX + 1) + 20
    mtg_start, _ = get_tuesday_meeting_window(sel_date, sel_date.year, sel_date.month)
    mtype         = tuesday_meeting_type(sel_date, sel_date.year, sel_date.month)
    mtg_label     = ("All Staff Meeting 11:00-12:00" if mtype == "all_staff"
                     else "CoP Meeting 11:00-12:00") if mtype else None

    # Build CSS using plain string concatenation — no nested bracket risk
    tw  = str(total_w); tc = str(TIME_COL); ic = str(INST_COL); rp = str(ROW_PX)
    css = ("<style>"
           + "*{box-sizing:border-box;margin:0;padding:0;font-family:-apple-system,sans-serif}"
           + "body{background:transparent}"
           + ".wrap{width:" + tw + "px}"
           + ".hdr-row{display:flex;margin-bottom:6px;padding-left:" + tc + "px}"
           + ".ih{width:" + ic + "px;flex-shrink:0;padding:0 3px}"
           + ".ih-inner{border-radius:6px;padding:7px 6px;text-align:center;"
           +          "font-weight:bold;font-size:12px;color:#fff;line-height:1.35}"
           + ".body{display:flex}"
           + ".tc{width:" + tc + "px;flex-shrink:0}"
           + ".ts{height:" + rp + "px;display:flex;align-items:center;justify-content:center;"
           +     "font-size:11px;font-weight:600;color:#64748b;background:#f1f5f9;"
           +     "border:1px solid #e2e8f0;border-radius:3px;margin-bottom:1px}"
           + ".ic{width:" + ic + "px;flex-shrink:0;position:relative;padding:0 3px}"
           + ".bg{height:" + rp + "px;border:1px solid #f0f0f0;border-radius:2px;margin-bottom:1px}"
           + ".bg-on{background:#fff}.bg-off{background:#f8fafc}"
           + ".bg-mtg{background:#fff3cd;border:1px solid #f59e0b}"
           + ".sb{position:absolute;left:3px;right:3px;border-radius:5px;"
           +     "padding:5px 30px 5px 8px;font-size:11px;color:#fff;line-height:1.4;"
           +     "overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.18)}"
           + ".bc{font-weight:700;font-size:11.5px}"
           + ".bt{opacity:.88;font-size:10.5px}.br{opacity:.80;font-size:10px}"
           + ".pen{position:absolute;top:5px;right:5px;background:rgba(255,255,255,.25);"
           +     "border:none;border-radius:4px;cursor:pointer;font-size:14px;"
           +     "padding:2px 5px;color:#fff}"
           + ".pen:hover{background:rgba(255,255,255,.45)}"
           + ".mb{position:absolute;left:3px;right:3px;background:#fef3c7;"
           +    "border:2px solid #f59e0b;border-radius:5px;display:flex;"
           +    "align-items:center;justify-content:center;"
           +    "font-size:11px;font-weight:700;color:#92400e;text-align:center}"
           + "</style>")

    h = ["<!DOCTYPE html><html><head><meta charset='utf-8'>" + css + "</head><body>"]
    h.append("<div class='wrap'><div class='hdr-row'>")
    for iname, sk in inst_list:
        ic_data = inst_color(iname)
        ic_bg   = ic_data["bg"]
        shlabel = DEFAULT_SHIFTS[sk]["label"]
        h.append("<div class='ih'><div class='ih-inner' style='background:" + ic_bg + "'>"
                 + iname
                 + "<br><span style='font-size:10px;opacity:.85'>"
                 + shlabel + "</span></div></div>")
    h.append("</div><div class='body'><div class='tc'>")
    for slot in slots:
        h.append("<div class='ts'>" + minutes_to_time(slot) + "</div>")
    h.append("</div>")

    for iname, sk in inst_list:
        sh_s4, sh_e4 = get_shift_window(sk)
        col_sess = sorted([s for s in day_sess if s["instructor"] == iname],
                          key=lambda x: x["class_start_min"])
        h.append("<div class='ic'>")
        for slot in slots:
            mtg_11 = time_to_minutes("11:00")
            mtg_12 = time_to_minutes("12:00")
            if (mtg_start is not None
                    and mtg_11 <= slot < mtg_12
                    and sh_s4 <= slot < sh_e4):
                bg_cls = "bg bg-mtg"
            elif sh_s4 <= slot < sh_e4:
                bg_cls = "bg bg-on"
            else:
                bg_cls = "bg bg-off"
            h.append("<div class='" + bg_cls + "'></div>")
        if mtg_label and mtg_start is not None and sh_s4 <= mtg_start < sh_e4:
            mtop = str((mtg_start - t_start) // SLOT_MINS * (ROW_PX + 1))
            mh   = str((time_to_minutes("12:00") - mtg_start) // SLOT_MINS * (ROW_PX + 1) - 3)
            h.append("<div class='mb' style='top:" + mtop + "px;height:" + mh + "px'>"
                     + "📋 " + mtg_label + "</div>")
        for s in col_sess:
            si2   = sess_idx.get(id(s), -1)
            s_col = inst_color(iname)
            bg    = "#dc2626" if si2 in conflicts else s_col["bg"]
            shad  = " [Shadow]" if s.get("shadow_of") else ""
            warn  = "⚠️ " if si2 in conflicts else ""
            top_px = str((s["class_start_min"] - t_start) // SLOT_MINS * (ROW_PX + 1))
            ht_px  = str(max(1, (s["class_end_min"] - s["class_start_min"])
                            // SLOT_MINS) * (ROW_PX + 1) - 3)
            h.append("<div class='sb' style='background:" + bg
                     + ";top:" + top_px + "px;height:" + ht_px + "px'>"
                     + "<div class='bc'>" + warn + s["course"] + shad + "</div>"
                     + "<div class='bt'>" + s["class_start"] + "–" + s["class_end"] + "</div>"
                     + "<div class='br'>" + s["room"] + "</div>"
                     + "<button class='pen' onclick='doEdit(" + str(si2) + ")'>✏️</button>"
                     + "</div>")
        h.append("</div>")

    h.append("</div></div>")
    h.append(
        "<script>"
        "function doEdit(idx){"
        "var u=new URL(window.parent.location.href);"
        "u.searchParams.set('_edit_idx',idx);"
        "window.parent.history.pushState({},'',u.toString());"
        "window.parent.dispatchEvent(new PopStateEvent('popstate',{state:{}}));}"
        "</script></body></html>")
    return "".join(h), total_h


def step9():
    import streamlit.components.v1 as components
    _init_edited()
    st.title("✏️ Edit Schedule")
    if not st.session_state.edited_schedule:
        st.warning("Please generate a schedule first (Step 5).")
        return
    sessions  = st.session_state.edited_schedule["sessions"]
    month     = st.session_state.schedule_month
    year      = st.session_state.schedule_year
    conflicts = _detect_conflicts(sessions)

    # URL bridge — pencil buttons set ?_edit_idx=N
    raw_idx = st.query_params.get("_edit_idx", None)
    if raw_idx is not None:
        try:
            sidx = int(raw_idx)
            if 0 <= sidx < len(sessions):
                st.session_state.edit_selected_idx = sidx
                st.session_state.edit_panel_mode   = "edit"
        except ValueError:
            pass
        st.query_params.clear()
        st.rerun()

    if (st.session_state.edit_panel_mode == "edit"
            and st.session_state.edit_selected_idx is not None
            and st.session_state.edit_selected_idx < len(sessions)):
        _edit_dialog(st.session_state.edit_selected_idx, sessions, month, year)

    if st.session_state.edit_panel_mode == "add":
        _add_dialog(sessions, month, year,
                    default_date=st.session_state.edit_selected_date)

    all_working = sorted({
        d.isoformat() for d in get_month_dates(year, month)
        if any(day_name(d) in DEFAULT_SHIFTS[sk]["days"] for sk in ACTIVE_SHIFTS)})
    if st.session_state.edit_selected_date not in all_working:
        act_d = sorted({s["date"] for s in sessions})
        st.session_state.edit_selected_date = (
            act_d[0] if act_d else (all_working[0] if all_working else None))
    if not all_working:
        st.info("No working days found.")
        return

    cur_idx = (all_working.index(st.session_state.edit_selected_date)
               if st.session_state.edit_selected_date in all_working else 0)

    n1, n2, n3, n4 = st.columns([1, 4, 1, 2])
    if n1.button("◄", use_container_width=True):
        if cur_idx > 0:
            st.session_state.edit_selected_date = all_working[cur_idx - 1]
            st.rerun()
    sel = n2.selectbox("Day", all_working, index=cur_idx,
        format_func=lambda x: date.fromisoformat(x).strftime("%A, %B %d"),
        label_visibility="collapsed")
    if sel != st.session_state.edit_selected_date:
        st.session_state.edit_selected_date = sel
        st.rerun()
    if n3.button("►", use_container_width=True):
        if cur_idx < len(all_working) - 1:
            st.session_state.edit_selected_date = all_working[cur_idx + 1]
            st.rerun()
    if n4.button("➕ Add New Class", use_container_width=True, type="primary"):
        st.session_state.edit_panel_mode   = "add"
        st.session_state.edit_selected_idx = None
        st.rerun()

    if conflicts:
        st.markdown(
            "<div style='background:#fef9c3;border:1px solid #ca8a04;border-radius:6px;"
            "padding:6px 12px;font-size:.85em;margin:6px 0'>"
            "⚠️ <b>" + str(len(conflicts)) + " conflict(s)</b>"
            " — shown in red. Click ✏️ to fix.</div>",
            unsafe_allow_html=True)

    st.markdown("---")
    sel_obj = date.fromisoformat(st.session_state.edit_selected_date)
    mtype_e = tuesday_meeting_type(sel_obj, sel_obj.year, sel_obj.month)
    mtg_tag = (" 📋 All Staff Meeting" if mtype_e == "all_staff"
               else (" 📋 CoP Meeting" if mtype_e == "cop" else ""))
    st.markdown(
        "<h3 style='margin-bottom:10px'>"
        + sel_obj.strftime("%A, %B") + " " + str(sel_obj.day)
        + ", " + str(sel_obj.year) + mtg_tag + "</h3>",
        unsafe_allow_html=True)

    html_out, est_h = _build_timeline_html(
        st.session_state.edit_selected_date, sessions, conflicts)
    components.html(html_out, height=est_h, scrolling=True)

    st.markdown("---")
    b1, b2 = st.columns(2)
    if b1.button("↩️ Reset to Generated Schedule", use_container_width=True):
        st.session_state.edited_schedule   = None
        st.session_state.edit_panel_mode   = None
        st.session_state.edit_selected_idx = None
        _init_edited()
        st.rerun()
    if EXCEL_OK:
        import io as _io
        buf = _io.BytesIO()
        build_excel(st.session_state.edited_schedule, month, year).save(buf)
        buf.seek(0)
        b2.download_button(
            "📥 Download Edited Excel", buf,
            file_name=calendar.month_name[month] + str(year) + "_EditedSchedule.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, key="dl_edited")


step = st.session_state.step
if   step == 1: step1()
elif step == 2: step2()
elif step == 3: step3()
elif step == 4: step4()
elif step == 5: step5()
elif step == 6: step6()
elif step == 7: step7()
elif step == 8: step8()
elif step == 9: step9()
